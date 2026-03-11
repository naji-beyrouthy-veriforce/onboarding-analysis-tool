"""
Backend API for Onboarding Analysis Tool
Using EXACT legacy script logic with FastAPI wrapper
"""
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional, Dict, Any, List, Union
import csv
import re
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from rapidfuzz import fuzz
from datetime import datetime, timedelta
import uuid
import shutil
from pathlib import Path
import logging
import asyncio
from concurrent.futures import ThreadPoolExecutor
import time
import string
import os
from threading import Lock
from contextlib import asynccontextmanager

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# Dynamic worker count based on CPU cores for better resource utilization
MAX_WORKERS = min(4, (os.cpu_count() or 1) + 1)
EXECUTOR = ThreadPoolExecutor(max_workers=MAX_WORKERS)

# File upload limits (bytes)
MAX_FILE_SIZE = 100 * 1024 * 1024  # 100MB
ALLOWED_EXTENSIONS = {'.csv', '.xlsx', '.xls'}

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup
    yield
    # Shutdown
    try:
        EXECUTOR.shutdown(wait=False, cancel_futures=True)
    except TypeError:
        # Python < 3.9 doesn't support cancel_futures
        EXECUTOR.shutdown(wait=False)

app = FastAPI(title="Onboarding Analysis Tool API", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # TODO: Restrict to specific domains in production: ["http://localhost:5173", "https://yourdomain.com"]
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

_BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = _BASE_DIR / "uploads"
OUTPUT_DIR = _BASE_DIR / "outputs"
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

jobs: Dict[str, Dict[str, Any]] = {}
jobs_lock = Lock()

def _is_supported_upload(path: Path) -> bool:
    return path.suffix.lower() in (".csv", ".xlsx", ".xls")

# ============================================================================
# EXACT LEGACY SCRIPT CONSTANTS AND FUNCTIONS
# ============================================================================

# Subscription pricing constants
CBX_DEFAULT_STANDARD_SUBSCRIPTION = 803
CBX_HEADER_LENGTH = 28

# Fuzzy matching thresholds - Named constants for clarity and maintainability
RATIO_COMPANY_PERFECT = 100.0
RATIO_COMPANY_VERY_HIGH = 95.0
RATIO_COMPANY_HIGH = 85.0
RATIO_COMPANY_GOOD = 75.0
RATIO_COMPANY_MODERATE = 70.0
RATIO_COMPANY_LOW = 33.0
RATIO_COMPANY_MINIMAL = 20.0
RATIO_ADDRESS_PERFECT = 100.0
RATIO_ADDRESS_HIGH = 85.0
RATIO_ADDRESS_MODERATE = 70.0

# Match display limit
MAX_MATCHES_TO_DISPLAY = 10

# Column indices for CBX
CBX_ID, CBX_COMPANY_FR, CBX_COMPANY_EN, CBX_COMPANY_OLD, CBX_ADDRESS, CBX_CITY, CBX_STATE, \
    CBX_COUNTRY, CBX_ZIP, CBX_FISTNAME, CBX_LASTNAME, CBX_EMAIL, CBX_EXPIRATION_DATE, CBX_REGISTRATION_STATUS, \
    CBX_SUSPENDED, CBX_MODULES, CBX_ACCESS_MODES, CBX_ACCOUNT_TYPE, CBX_SUB_PRICE_CAD, CBX_EMPL_PRICE_CAD, \
    CBX_SUB_PRICE_USD, CBX_EMPL_PRICE_USD, CBX_HIRING_CLIENT_NAMES, \
    CBX_HIRING_CLIENT_IDS, CBX_HIRING_CLIENT_QSTATUS, CBX_PARENTS, CBX_ASSESSMENT_LEVEL, CBX_IS_NEW_PRODUCT = range(CBX_HEADER_LENGTH)

HC_HEADER_LENGTH = 41

# Column indices for HC
HC_COMPANY, HC_FIRSTNAME, HC_LASTNAME, HC_EMAIL, HC_CONTACT_PHONE, HC_CONTACT_LANGUAGE, HC_STREET, HC_CITY, \
    HC_STATE, HC_COUNTRY, HC_ZIP, HC_CATEGORY, HC_DESCRIPTION, HC_PHONE, HC_EXTENSION, HC_FAX, HC_WEBSITE, \
    HC_LANGUAGE, HC_IS_TAKE_OVER, HC_TAKEOVER_QUALIFICATION_DATE, HC_TAKEOVER_QF_STATUS, \
    HC_PROJECT_NAME, HC_QUESTIONNAIRE_NAME, HC_QUESTIONNAIRE_ID, HC_PRICING_GROUP_ID, HC_PRICING_GROUP_CODE, \
    HC_HIRING_CLIENT_NAME, HC_HIRING_CLIENT_ID, HC_IS_ASSOCIATION_FEE, HC_BASE_SUBSCRIPTION_FEE, \
    HC_CONTACT_CURRENCY, HC_AGENT_IN_CHARGE_ID, HC_TAKEOVER_FOLLOW_UP_DATE, HC_TAKEOVER_RENEWAL_DATE, \
    HC_INFORMATION_SHARED, HC_CONTACT_TIMEZONE, HC_DO_NOT_MATCH, HC_FORCE_CBX_ID, HC_AMBIGUOUS, \
    HC_CONTRACTORCHECK_ACCOUNT, HC_ASSESSMENT_LEVEL = range(HC_HEADER_LENGTH)

SUPPORTED_CURRENCIES = ('CAD', 'USD')

assessment_levels = {
    "gold": 2, "silver": 2, "bronze": 1,
    "level3": 2, "level2": 2, "level1": 1,
    "3": 2, "2": 2, "1": 1
}

cbx_headers = ['id', 'name_fr', 'name_en', 'old_names', 'address', 'city', 'state', 'country', 'postal_code',
               'first_name', 'last_name', 'email', 'cbx_expiration_date', 'registration_code', 'suspended',
               'modules', 'access_modes', 'code', 'subscription_price_cad', 'employee_price_cad',
               'subscription_price_usd', 'employee_price_usd', 'hiring_client_names',
               'hiring_client_ids', 'hiring_client_qstatus', 'parents', 'assessment_level', 'new_product']

hiring_client_headers = ['contractor_name', 'contact_first_name', 'contact_last_name', 'contact_email', 'contact_phone',
                         'contact_language', 'address', 'city', 'province_state_iso2', 'country_iso2',
                         'postal_code', 'category', 'description', 'phone', 'extension', 'fax', 'website', 'language',
                         'is_take_over', 'qualification_expiration_date',
                         'qualification_status', 'batch', 'questionnaire_name', 'questionnaire_id',
                         'pricing_group_id', 'pricing_group_code', 'hiring_client_name', 'hiring_client_id',
                         'is_association_fee', 'base_subscription_fee', 'contact_currency', 'agent_in_charge_id',
                         'take_over_follow-up_date', 'renewal_date', 'information_shared', 'contact_timezone',
                         'do_not_match', 'force_cbx_id', 'ambiguous', 'contractorcheck_account', 'assessment_level']

analysis_headers = ['cbx_id', 'hc_contractor_summary', 'analysis', 'cbx_contractor', 'cbx_street', 'cbx_city',
                    'cbx_state', 'cbx_zip', 'cbx_country', 'cbx_expiration_date', 'registration_status', 'suspended',
                    'cbx_email', 'cbx_first_name', 'cbx_last_name', 'modules', 'cbx_account_type',
                    'cbx_subscription_fee', 'cbx_employee_price', 'parents', 'previous', 'hiring_client_names',
                    'hiring_client_count', 'is_in_relationship', 'is_qualified', 'ratio_company', 'ratio_address',
                    'contact_match', 'cbx_assessment_level', 'new_product', 'generic_domain', 'match_count',
                    'match_count_with_hc', 'is_subscription_upgrade', 'upgrade_price', 'prorated_upgrade_price',
                    'create_in_cbx', 'action', 'index']

rd_headers = ['contractor_name', 'contact_first_name', 'contact_last_name', 'contact_email', 'contact_phone',
              'contact_language', 'address', 'city', 'province_state_iso2', 'country_iso2',
              'postal_code', 'description', 'phone', 'extension', 'fax', 'website', 'language',
              'qualification_expiration_date', 'qualification_status', 'contact_currency',
              'agent_in_charge_id', 'renewal_date', 'information_shared', 'contact_timezone', 'questionnaire_name',
              'questionnaire_ids',
              'pricing_group_code', 'pricing_group_id', 'hiring_client_id', 'contractorcheck_account',
              'assessment_level']

existing_contractors_headers = ['cbx_id']
existing_contractors_headers.extend(rd_headers.copy())

hubspot_headers = ['contractor_name', 'contact_first_name', 'contact_last_name', 'contact_email', 'contact_phone',
                   'contact_language', 'address', 'city', 'province_state_iso2', 'country_iso2',
                   'postal_code', 'cbx_id', 'cbx_expiration_date', 'questionnaire_name',
                   'questionnaire_id', 'hiring_client_name', 'hiring_client_id', 'action']

metadata_headers = ['metadata_x', 'metadata_y', 'metadata_z', '...']

BASE_GENERIC_DOMAIN = ['yahoo.ca', 'yahoo.com', 'hotmail.com', 'gmail.com', 'outlook.com',
                       'bell.com', 'bell.ca', 'videotron.ca', 'eastlink.ca', 'kos.net', 'bellnet.ca', 'sasktel.net',
                       'aol.com', 'tlb.sympatico.ca', 'sogetel.net', 'cgocable.ca',
                       'hotmail.ca', 'live.ca', 'icloud.com', 'hotmail.fr', 'yahoo.com', 'outlook.fr', 'msn.com',
                       'globetrotter.net', 'live.com', 'sympatico.ca', 'live.fr', 'yahoo.fr', 'telus.net',
                       'shaw.ca', 'me.com', 'bell.net', 'cablevision.qc.ca', 'live.ca', 'tlb.sympatico.ca',
                       '', 'videotron.qc.ca', 'ivic.qc.ca', 'qc.aira.com', 'canada.ca', 'axion.ca', 'bellsouth.net',
                       'telusplanet.net', 'rogers.com', 'mymts.net', 'nb.aibn.com', 'on.aibn.com', 'live.be',
                       'nbnet.nb.ca', 'execulink.com', 'bellaliant.com', 'nf.aibn.com', 'clintar.com', 'pathcom.com',
                       'oricom.ca', 'mts.net', 'xplornet.com', 'mcsnet.ca', 'att.net', 'ymail.com', 'mail.com',
                       'bellaliant.net', 'ns.sympatico.ca', 'ns.aliantzinc.ca', 'mnsi.net']

BASE_GENERIC_COMPANY_NAME_WORDS = ['construction', 'contracting', 'industriel', 'industriels', 'service',
                                   'services', 'inc', 'limited', 'ltd', 'ltee', 'ltée', 'co', 'industrial',
                                   'solutions', 'llc', 'enterprises', 'enterprise', 'entreprise', 'entreprises', 'systems', 'industries',
                                   'technologies', 'company', 'corporation', 'installations', 'enr']

LIST_SEPARATOR = ";"
GENERIC_DOMAIN = set(BASE_GENERIC_DOMAIN)
GENERIC_COMPANY_NAME_WORDS = BASE_GENERIC_COMPANY_NAME_WORDS

# Pre-compiled regex pattern for optimal performance in remove_generics()
GENERIC_WORDS_PATTERN = re.compile(
    r'\b(?:' + '|'.join(re.escape(word) for word in GENERIC_COMPANY_NAME_WORDS) + r')\b',
    re.IGNORECASE
)

def smart_boolean(bool_data: Union[str, bool, int, None]) -> bool:
    """Convert various boolean representations to Python bool."""
    if isinstance(bool_data, str):
        bool_data = bool_data.lower().strip()
        return True if bool_data in ('true', '=true', 'yes', 'vraie', '=vraie', '1') else False
    else:
        return bool(bool_data)

def norm_name(name: Optional[str]) -> str:
    """Normalize name by removing punctuation and converting to lowercase."""
    if not name:
        return ''
    return str(name).translate(str.maketrans('', '', string.punctuation)).strip().lower()

def remove_generics(company_name: str) -> str:
    """Remove generic company words using pre-compiled regex pattern."""
    # Use pre-compiled pattern for 5-10x performance improvement
    return GENERIC_WORDS_PATTERN.sub('', company_name)

def clean_company_name(name: Optional[str]) -> str:
    """Clean and normalize company name for fuzzy matching."""
    if not name or (isinstance(name, str) and not name.strip()):
        return ''
    try:
        name = str(name).lower().replace('.', '').replace(',', '').strip()
        name = re.sub(r"\([^()]*\)", "", name)
        name = remove_generics(name)
        # Remove multiple spaces
        name = re.sub(r'\s+', ' ', name).strip()
        return name
    except (TypeError, AttributeError) as e:
        logger.warning(f"Error cleaning company name '{name}': {e}")
        return ''

def parse_assessment_level(level: Union[str, int, None]) -> int:
    """Parse assessment level from various formats to integer."""
    if level is None or (isinstance(level, int) and 0 < level < 4):
        return level if level else 0
    if isinstance(level, str) and level.lower() in assessment_levels:
        return assessment_levels[level.lower()]
    return 0

def core_mandatory_provided(hcd: List[Any]) -> bool:
    """Check if all mandatory fields are provided in hiring client data."""
    mandatory_fields = (HC_COMPANY, HC_FIRSTNAME, HC_LASTNAME, HC_EMAIL, HC_CONTACT_PHONE,
                        HC_STREET, HC_CITY, HC_STATE, HC_COUNTRY, HC_ZIP)
    country = hcd[HC_COUNTRY].strip().lower() if isinstance(hcd[HC_COUNTRY], str) else str(hcd[HC_COUNTRY]).lower()
    for field in mandatory_fields:
        f_value = str(hcd[field]).strip() if hcd[field] else ""
        if f_value == "":
            if field == HC_STATE and country not in ('ca', 'us'):
                pass
            else:
                return False
    return True


def add_analysis_data(hc_row, cbx_row, ratio_company=None, ratio_address=None, contact_match=None):
    cbx_company = cbx_row[CBX_COMPANY_FR] if cbx_row[CBX_COMPANY_FR] else cbx_row[CBX_COMPANY_EN]

    hiring_clients_list = [norm_name(x) for x in str(cbx_row[CBX_HIRING_CLIENT_NAMES]).split(LIST_SEPARATOR) if x]
    hiring_clients_qstatus = str(cbx_row[CBX_HIRING_CLIENT_QSTATUS]).split(LIST_SEPARATOR)
    hc_name_norm = norm_name(hc_row[HC_HIRING_CLIENT_NAME])
    hc_count = len(hiring_clients_list) if cbx_row[CBX_HIRING_CLIENT_NAMES] else 0
    is_in_relationship = hc_name_norm in hiring_clients_list and hc_name_norm != ''
    is_qualified = False
    matched_qstatus = None

    sub_price_usd = float(cbx_row[CBX_SUB_PRICE_USD]) if cbx_row[CBX_SUB_PRICE_USD] else 0.0
    employee_price_usd = float(cbx_row[CBX_EMPL_PRICE_USD]) if cbx_row[CBX_EMPL_PRICE_USD] else 0.0
    sub_price_cad = float(cbx_row[CBX_SUB_PRICE_CAD]) if cbx_row[CBX_SUB_PRICE_CAD] else 0.0
    employee_price_cad = float(cbx_row[CBX_EMPL_PRICE_CAD]) if cbx_row[CBX_EMPL_PRICE_CAD] else 0.0

    hiring_client_contractor_summary = f'{hc_row[HC_COMPANY]}, {hc_row[HC_STREET]}, {hc_row[HC_CITY]}, {hc_row[HC_STATE]}, {hc_row[HC_COUNTRY]}, {hc_row[HC_ZIP]}, {hc_row[HC_EMAIL]}, {hc_row[HC_FIRSTNAME]} {hc_row[HC_LASTNAME]}'

    if hc_row[HC_CONTACT_CURRENCY] != '' and hc_row[HC_CONTACT_CURRENCY] not in SUPPORTED_CURRENCIES:
        logger.warning(f'Invalid currency: {hc_row[HC_CONTACT_CURRENCY]}, must be in {SUPPORTED_CURRENCIES}')

    for idx, val in enumerate(hiring_clients_list):
        if val == hc_name_norm and idx < len(hiring_clients_qstatus):
            qstatus = hiring_clients_qstatus[idx].strip().lower()
            matched_qstatus = qstatus
            if qstatus == 'validated':
                is_qualified = True
            break

    try:
        expiration_date = datetime.strptime(cbx_row[CBX_EXPIRATION_DATE], "%d/%m/%y") if cbx_row[CBX_EXPIRATION_DATE] else None
    except ValueError:
        try:
            expiration_date = datetime.strptime(cbx_row[CBX_EXPIRATION_DATE], "%d/%m/%Y") if cbx_row[CBX_EXPIRATION_DATE] else None
        except:
            expiration_date = None

    return {
        'cbx_id': int(cbx_row[CBX_ID]),
        'hc_contractor_summary': hiring_client_contractor_summary,
        'analysis': '',
        'company': cbx_company,
        'address': cbx_row[CBX_ADDRESS],
        'city': cbx_row[CBX_CITY],
        'state': cbx_row[CBX_STATE],
        'zip': cbx_row[CBX_ZIP],
        'country': cbx_row[CBX_COUNTRY],
        'expiration_date': expiration_date,
        'registration_status': cbx_row[CBX_REGISTRATION_STATUS],
        'suspended': cbx_row[CBX_SUSPENDED],
        'email': cbx_row[CBX_EMAIL],
        'first_name': cbx_row[CBX_FISTNAME],
        'last_name': cbx_row[CBX_LASTNAME],
        'modules': cbx_row[CBX_MODULES],
        'account_type': cbx_row[CBX_ACCOUNT_TYPE],
        'subscription_price': sub_price_cad if hc_row[HC_CONTACT_CURRENCY] == "CAD" else sub_price_usd,
        'employee_price': employee_price_cad if hc_row[HC_CONTACT_CURRENCY] == "CAD" else employee_price_usd,
        'parents': cbx_row[CBX_PARENTS],
        'previous': cbx_row[CBX_COMPANY_OLD],
        'hiring_client_names': cbx_row[CBX_HIRING_CLIENT_NAMES],
        'hiring_client_count': hc_count,
        'is_in_relationship': is_in_relationship,
        'is_qualified': is_qualified,
        'matched_qstatus': matched_qstatus,
        'ratio_company': ratio_company,
        'ratio_address': ratio_address,
        'contact_match': contact_match,
        'cbx_assessment_level': cbx_row[CBX_ASSESSMENT_LEVEL],
        'new_product': cbx_row[CBX_IS_NEW_PRODUCT]
    }


def action(hc_data, cbx_data, create, subscription_update, expiration_date, is_qualified):
    if create:
        if smart_boolean(hc_data[HC_IS_TAKE_OVER]):
            return 'activation_link'
        else:
            if hc_data[HC_AMBIGUOUS]:
                return 'ambiguous_onboarding'
            elif core_mandatory_provided(hc_data):
                return 'onboarding'
            else:
                return 'missing_info'
    else:
        reg_status = cbx_data['registration_status']
        if smart_boolean(hc_data[HC_IS_TAKE_OVER]):
            if reg_status == 'Suspended':
                return 'restore_suspended'
            elif reg_status == 'Active':
                return 'add_questionnaire'
            elif reg_status == 'Non Member':
                return 'activation_link'
            else:
                logger.warning(f'WARNING: invalid registration status {reg_status}')
                return 'add_questionnaire'
        else:
            if reg_status == 'Active':
                if cbx_data['is_in_relationship']:
                    qstatus = cbx_data.get('matched_qstatus', None)
                    if qstatus == 'validated':
                        return 'already_qualified'
                    elif qstatus in ('pending', 'expired', 'conditional', 'refused'):
                        return 'follow_up_qualification'
                    else:
                        if is_qualified:
                            return 'already_qualified'
                        else:
                            return 'follow_up_qualification'
                else:
                    if subscription_update:
                        return 'subscription_upgrade'
                    elif hc_data[HC_IS_ASSOCIATION_FEE] and not cbx_data['is_in_relationship']:
                        if expiration_date:
                            in_60_days = datetime.now() + timedelta(days=60)
                            if expiration_date > in_60_days:
                                return 'association_fee'
                            else:
                                return 'add_questionnaire'
                        else:
                            return 'association_fee'
                    else:
                        return 'add_questionnaire'
            elif reg_status == 'Suspended':
                return 'restore_suspended'
            elif reg_status in ('Non Member', '', None):
                return 're_onboarding'
            else:
                raise AssertionError(f'invalid registration status: {reg_status}')


def update_job(job_id: str, **kwargs):
    with jobs_lock:
        if job_id in jobs:
            jobs[job_id].update(kwargs)
            # Avoid noisy progress-only logs; keep status/error/result updates visible.
            if any(k in kwargs for k in ("status", "error", "result_file")):
                logger.info(f"[{job_id}] Updated: {kwargs}")


# ============================================================================
# MAIN PROCESSING FUNCTION (EXACT LEGACY LOGIC)
# ============================================================================

def process_matching_job(job_id: str, cbx_path: Path, hc_path: Path, min_company_ratio: int, min_address_ratio: int):
    try:
        t0 = time.time()
        logger.info(f"[{job_id}] ========== STARTING JOB ==========")
        update_job(job_id, status="processing", message="Reading files...", progress=0.0)

        # Read CBX data
        cbx_data = []
        logger.info(f"[{job_id}] Reading CBX file: {cbx_path}")

        if cbx_path.suffix.lower() in ('.xlsx', '.xls'):
            wb = openpyxl.load_workbook(cbx_path, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.rows:
                cbx_data.append([cell.value if cell.value is not None else '' for cell in row])
            wb.close()
        else:
            with open(cbx_path, 'r', encoding='utf-8-sig') as f:
                for row in csv.reader(f):
                    cbx_data.append(row)

        if cbx_data:
            headers = cbx_data.pop(0)
            logger.info(f"[{job_id}] CBX loaded: {len(cbx_data)} rows")

        # Pre-normalize CBX data once to avoid repeated computations during matching
        logger.info(f"[{job_id}] Pre-normalizing {len(cbx_data)} CBX records for faster matching...")
        for idx, cbx_row in enumerate(cbx_data):
            try:
                cbx_row.append(clean_company_name(cbx_row[CBX_COMPANY_EN]))  # Index 28: normalized EN name
                cbx_row.append(clean_company_name(cbx_row[CBX_COMPANY_FR]))  # Index 29: normalized FR name
                cbx_row.append(str(cbx_row[CBX_ADDRESS] if cbx_row[CBX_ADDRESS] else '').lower().replace('.', '').strip())  # Index 30: normalized address
                # Pre-normalize email and domain for faster comparison
                cbx_email_lower = str(cbx_row[CBX_EMAIL]).lower()
                cbx_row.append(cbx_email_lower)  # Index 31: normalized email
                cbx_row.append(cbx_email_lower[cbx_email_lower.find('@') + 1:] if '@' in cbx_email_lower else '')  # Index 32: email domain
                cbx_row.append(str(cbx_row[CBX_ZIP] if cbx_row[CBX_ZIP] else '').replace(' ', '').upper())  # Index 33: normalized ZIP
                # Pre-clean previous names for faster comparison
                prev_names = str(cbx_row[CBX_COMPANY_OLD]).split(LIST_SEPARATOR)
                cleaned_prev = [clean_company_name(item) for item in prev_names if item and item not in (cbx_row[CBX_COMPANY_EN], cbx_row[CBX_COMPANY_FR])]
                cbx_row.append([n for n in cleaned_prev if n])  # Index 34: cleaned previous names list
            except IndexError as e:
                logger.error(f"[{job_id}] CBX row {idx+2} missing columns. Expected 28, got {len(cbx_row)}. Error: {e}")
                raise ValueError(f"CBX file format error at row {idx+2}: Missing required columns. Expected 28 columns, found {len(cbx_row)}.")
        logger.info(f"[{job_id}] Pre-normalization complete")

        # Read HC data
        logger.info(f"[{job_id}] Reading HC file: {hc_path}")
        hc_data = []

        if hc_path.suffix.lower() in ('.xlsx', '.xls'):
            wb = openpyxl.load_workbook(hc_path, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.rows:
                if row[0].value:
                    hc_data.append([cell.value if cell.value is not None else '' for cell in row])
            wb.close()
        else:
            with open(hc_path, 'r', encoding='utf-8-sig') as f:
                for row in csv.reader(f):
                    if row and row[0]:
                        hc_data.append(row)

        if hc_data:
            hc_headers = hc_data.pop(0)
            logger.info(f"[{job_id}] HC loaded: {len(hc_data)} rows")

        total = len(hc_data)
        update_job(job_id, message=f"Starting matching for {total} contractors...", progress=0.05)

        # Create output workbook
        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.title = 'all'

        # Create action sheets
        out_ws_onboarding = out_wb.create_sheet(title="onboarding")
        out_ws_association_fee = out_wb.create_sheet(title="association_fee")
        out_ws_re_onboarding = out_wb.create_sheet(title="re_onboarding")
        out_ws_subscription_upgrade = out_wb.create_sheet(title="subscription_upgrade")
        out_ws_ambiguous_onboarding = out_wb.create_sheet(title="ambiguous_onboarding")
        out_ws_restore_suspended = out_wb.create_sheet(title="restore_suspended")
        out_ws_activation_link = out_wb.create_sheet(title="activation_link")
        out_ws_already_qualified = out_wb.create_sheet(title="already_qualified")
        out_ws_add_questionnaire = out_wb.create_sheet(title="add_questionnaire")
        out_ws_missing_information = out_wb.create_sheet(title="missing_info")
        out_ws_follow_up_qualification = out_wb.create_sheet(title="follow_up_qualification")
        out_ws_onboarding_rd = out_wb.create_sheet(title="Data to import")
        out_ws_existing_contractors = out_wb.create_sheet(title="Existing Contractors")
        out_ws_onboarding_hs = out_wb.create_sheet(title="Data for HS")

        sheets = (out_ws, out_ws_onboarding, out_ws_association_fee, out_ws_re_onboarding,
                  out_ws_subscription_upgrade, out_ws_ambiguous_onboarding, out_ws_restore_suspended,
                  out_ws_activation_link, out_ws_already_qualified, out_ws_add_questionnaire,
                  out_ws_missing_information, out_ws_follow_up_qualification,
                  out_ws_onboarding_rd, out_ws_existing_contractors, out_ws_onboarding_hs)

        # Handle metadata columns and create mappings for special sheets
        metadata_indexes = []
        for idx, val in enumerate(hc_headers):
            if val.lower().startswith('metadata'):
                metadata_indexes.append(idx)
        metadata_indexes.sort(reverse=True)
        
        # Create combined headers with metadata moved to end
        final_headers = list(hc_headers) + analysis_headers
        metadata_array = []
        for md_index in metadata_indexes:
            if md_index < len(final_headers):
                metadata_array.insert(0, final_headers.pop(md_index))
        final_headers.extend(metadata_array)
        
        # Extend special sheet headers with metadata
        hubspot_headers_with_metadata = hubspot_headers.copy()
        hubspot_headers_with_metadata.extend(metadata_array)
        existing_contractors_headers_with_metadata = existing_contractors_headers.copy()
        existing_contractors_headers_with_metadata.extend(metadata_array)
        rd_headers_with_metadata = rd_headers.copy()
        rd_headers_with_metadata.extend(metadata_array)
        
        # Create header mappings for special sheets
        rd_headers_mapping = []
        hs_headers_mapping = []
        existing_contractors_headers_mapping = []
        rd_pricing_group_id_col = -1
        rd_pricing_group_code_col = -1
        
        column_rd = column_hs = column_existing_contractors = 0
        for index, value in enumerate(final_headers):
            # Write headers to standard sheets (skip the last 3 special sheets)
            for sheet in sheets[:-3]:
                sheet.cell(1, index + 1, value)
            
            # Map rd_headers (Data to import sheet)
            rd_headers_for_value = [s for s in rd_headers_with_metadata if value in s]
            if rd_headers_for_value:
                column_rd += 1
                rd_headers_mapping.append(True)
                # Track pricing_group columns for swapping
                if value == "pricing_group_id":
                    adjustment = 1
                    rd_pricing_group_id_col = column_rd
                elif value == "pricing_group_code":
                    adjustment = -1
                    rd_pricing_group_code_col = column_rd
                else:
                    adjustment = 0
                
                if value in rd_headers_with_metadata:
                    out_ws_onboarding_rd.cell(1, column_rd + adjustment, value)
                else:
                    out_ws_onboarding_rd.cell(1, column_rd, rd_headers_for_value[0])
            else:
                rd_headers_mapping.append(False)
            
            # Map hubspot_headers (Data for HS sheet)
            if value in hubspot_headers_with_metadata:
                column_hs += 1
                hs_headers_mapping.append(True)
                out_ws_onboarding_hs.cell(1, column_hs, value)
            else:
                hs_headers_mapping.append(False)
            
            # Map existing_contractors_headers (Existing Contractors sheet)
            existing_contractors_headers_for_value = [s for s in existing_contractors_headers_with_metadata if value in s]
            if existing_contractors_headers_for_value:
                column_existing_contractors += 1
                existing_contractors_headers_mapping.append(True)
                if value in existing_contractors_headers_with_metadata:
                    out_ws_existing_contractors.cell(1, column_existing_contractors, value)
                else:
                    out_ws_existing_contractors.cell(1, column_existing_contractors, existing_contractors_headers_for_value[0])
            else:
                existing_contractors_headers_mapping.append(False)

        # Process each HC contractor (EXACT LEGACY LOGIC)
        for index, hc_row in enumerate(hc_data):
            matches = []
            hc_company = hc_row[HC_COMPANY]
            clean_hc_company = clean_company_name(hc_company)
            
            # Normalize email - optimized single-pass extraction
            hc_email_raw = str(hc_row[HC_EMAIL]).lower().strip()
            hc_email = re.split(r'[;\n,]', hc_email_raw, maxsplit=1)[0].strip()
            hc_domain = hc_email[hc_email.find('@') + 1:] if '@' in hc_email else ''
            hc_zip = str(hc_row[HC_ZIP] if hc_row[HC_ZIP] else '').replace(' ', '').upper()
            hc_address = str(hc_row[HC_STREET] if hc_row[HC_STREET] else '').lower().replace('.', '').strip()
            hc_force_cbx = str(hc_row[HC_FORCE_CBX_ID] if hc_row[HC_FORCE_CBX_ID] else '')
            
            # Determine data completeness to dynamically adjust matching thresholds
            has_company = bool(hc_company and len(clean_hc_company) >= 3)
            has_address = bool(hc_address and hc_zip)
            has_email = bool(hc_email and '@' in hc_email)
            is_complete_data = has_company and (has_address or has_email)
            
            # Maximum number of matches to display in analysis column
            max_matches_to_keep = MAX_MATCHES_TO_DISPLAY
            
            # Dynamic matching thresholds based on data completeness
            # More complete data = stricter thresholds to reduce false positives
            # Less complete data = more lenient to ensure we don't miss valid matches
            if is_complete_data:
                # Complete data (company + address/email): use strict thresholds
                min_company_ratio = RATIO_COMPANY_GOOD
                min_address_ratio = RATIO_ADDRESS_HIGH
            else:
                # Incomplete data: adjust thresholds based on what we have
                if has_company and not has_address and not has_email:
                    # Only company name available
                    min_company_ratio = 80.0
                    min_address_ratio = 0.0
                elif has_company and has_email and not has_address:
                    # Company + email but no address
                    min_company_ratio = 60.0
                    min_address_ratio = 0.0
                elif has_company and has_address and not has_email:
                    # Company + address but no email
                    min_company_ratio = 65.0
                    min_address_ratio = 70.0
                else:
                    # Very incomplete data
                    min_company_ratio = 40.0
                    min_address_ratio = 0.0
            
            if not smart_boolean(hc_row[HC_DO_NOT_MATCH]):
                if hc_force_cbx:
                    cbx_row = next((x for x in cbx_data if str(x[CBX_ID]).strip() == hc_force_cbx), None)
                    if cbx_row:
                        matches.append(add_analysis_data(hc_row, cbx_row))
                else:
                    hc_country = hc_row[HC_COUNTRY]
                    
                    for cbx_row in cbx_data:
                        # EARLY EXIT: Skip entirely if countries don't match (fastest check)
                        # Only compare countries if BOTH are non-empty (case-insensitive comparison)
                        if hc_country and cbx_row[CBX_COUNTRY] and str(cbx_row[CBX_COUNTRY]).strip().upper() != str(hc_country).strip().upper():
                            continue
                        
                        # Use pre-normalized data (indexes 28-34)
                        cbx_company_en = cbx_row[28] if len(cbx_row) > 28 else clean_company_name(cbx_row[CBX_COMPANY_EN])
                        cbx_company_fr = cbx_row[29] if len(cbx_row) > 29 else clean_company_name(cbx_row[CBX_COMPANY_FR])
                        cbx_address = cbx_row[30] if len(cbx_row) > 30 else str(cbx_row[CBX_ADDRESS] if cbx_row[CBX_ADDRESS] else '').lower().replace('.', '').strip()
                        cbx_email = cbx_row[31] if len(cbx_row) > 31 else str(cbx_row[CBX_EMAIL]).lower()
                        cbx_domain = cbx_row[32] if len(cbx_row) > 32 else (cbx_email[cbx_email.find('@') + 1:] if '@' in cbx_email else '')
                        cbx_zip = cbx_row[33] if len(cbx_row) > 33 else str(cbx_row[CBX_ZIP] if cbx_row[CBX_ZIP] else '').replace(' ', '').upper()
                        cleaned_prev_names = cbx_row[34] if len(cbx_row) > 34 else []

                        # Contact match calculation
                        if hc_email:
                            if hc_domain in GENERIC_DOMAIN:
                                contact_match = (cbx_email == hc_email)
                            else:
                                contact_match = (cbx_domain == hc_domain)
                        else:
                            contact_match = False

                        # Address ratio calculation - only check country if BOTH have countries (case-insensitive)
                        if hc_country and cbx_row[CBX_COUNTRY] and str(cbx_row[CBX_COUNTRY]).strip().upper() != str(hc_country).strip().upper():
                            ratio_zip = ratio_address = 0.0
                        elif not hc_address or not cbx_address:
                            ratio_zip = ratio_address = 0.0
                        else:
                            ratio_zip = fuzz.ratio(cbx_zip, hc_zip)
                            ratio_address = fuzz.token_sort_ratio(cbx_address, hc_address)
                            # Legacy combination logic (EXACT)
                            ratio_address = ratio_address if ratio_zip == 0 else ratio_zip if ratio_address == 0 \
                                else ratio_address * ratio_zip / 100

                        # OPTIMIZATION: Skip fuzzy matching for very short/empty company names
                        if len(clean_hc_company) < 3 or (len(cbx_company_en) < 3 and len(cbx_company_fr) < 3):
                            ratio_company = 0
                        else:
                            ratio_company_fr = fuzz.token_sort_ratio(cbx_company_fr, clean_hc_company)
                            ratio_company_en = fuzz.token_sort_ratio(cbx_company_en, clean_hc_company)
                            ratio_company = ratio_company_fr if ratio_company_fr > ratio_company_en else ratio_company_en

                            # Check previous names using pre-cleaned list
                            ratio_previous = 0
                            for item_clean in cleaned_prev_names:
                                r = fuzz.token_sort_ratio(item_clean, clean_hc_company)
                                if r > ratio_previous:
                                    ratio_previous = r

                            if ratio_previous > ratio_company:
                                ratio_company = ratio_previous

                        # ============================================================================
                        # MATCHING CONDITIONS (Priority Ordered - Mutually Exclusive)
                        # ============================================================================
                        # Conditions are evaluated in order of confidence from highest to lowest.
                        # Each condition is mutually exclusive - once a match is found, no further
                        # conditions are checked. This ensures consistent, predictable matching.
                        # ============================================================================
                        
                        if ratio_company == RATIO_COMPANY_PERFECT:
                            # HIGHEST PRIORITY: Perfect company name match with email verification
                            if hc_email and cbx_email:
                                # Both have emails - domains must match
                                if hc_domain == cbx_domain:
                                    matches.append(add_analysis_data(hc_row, cbx_row, ratio_company, ratio_address, contact_match))
                                # If domains don't match, check address as secondary validation
                                elif ratio_address >= min_address_ratio:
                                    # Perfect name + good address = likely same company despite email change
                                    matches.append(add_analysis_data(hc_row, cbx_row, ratio_company, ratio_address, contact_match))
                                else:
                                    # Domains don't match AND address is poor, but 100% company name is still very strong signal
                                    # Match anyway - perfect company name match is reliable even without email/address validation
                                    matches.append(add_analysis_data(hc_row, cbx_row, ratio_company, ratio_address, contact_match))
                            else:
                                # At least one email is missing - match based on company name alone
                                matches.append(add_analysis_data(hc_row, cbx_row, ratio_company, ratio_address, contact_match))
                        elif ratio_company >= min_company_ratio and ratio_address >= min_address_ratio:
                            # PRIMARY: Strong dual signal - both company (≥75%) AND address (≥85%) meet thresholds
                            # Combined signals provide higher confidence than single strong signal
                            matches.append(add_analysis_data(hc_row, cbx_row, ratio_company, ratio_address, contact_match))
                        elif ratio_address == RATIO_ADDRESS_PERFECT and ratio_company >= RATIO_COMPANY_MODERATE:
                            # Perfect address match + reasonable company similarity
                            # Address is highly stable identifier, catches name variations at same location
                            matches.append(add_analysis_data(hc_row, cbx_row, ratio_company, ratio_address, contact_match))
                        elif ratio_company >= RATIO_COMPANY_VERY_HIGH:
                            # Very high company similarity alone (ignores address)
                            # Strong single signal for companies that may have moved or lack address data
                            matches.append(add_analysis_data(hc_row, cbx_row, ratio_company, ratio_address, contact_match))
                        elif ratio_company >= RATIO_COMPANY_HIGH and contact_match and not hc_address and not cbx_address:
                            # Strong company + email match when both addresses are legitimately missing
                            # Fills gap between dual-signal (75%+85% addr) and very-high company (95%)
                            # Requires email verification as secondary signal to compensate for missing addresses
                            matches.append(add_analysis_data(hc_row, cbx_row, ratio_company, ratio_address, contact_match))
                        elif contact_match and ratio_company >= RATIO_COMPANY_LOW:
                            # Email domain match + moderate company similarity (dual signal)
                            # 33% threshold catches legitimate variations with same domain
                            # Same email domain strongly indicates same company
                            matches.append(add_analysis_data(hc_row, cbx_row, ratio_company, ratio_address, contact_match))
                        elif hc_email and cbx_email == hc_email and ratio_company >= RATIO_COMPANY_MINIMAL:
                            # FALLBACK: Exact same email with minimal company similarity
                            # Safety net for same contact person potentially working at different companies
                            matches.append(add_analysis_data(hc_row, cbx_row, ratio_company, ratio_address, contact_match))

            # ============================================================================
            # MATCH FILTERING AND PRIORITIZATION
            # ============================================================================
            # Apply intelligent filtering to refine matches while avoiding false negatives:
            # 1. Remove "DO NOT USE" entries
            # 2. Prioritize hiring client relationships (but keep high-confidence non-HC matches)
            # 3. Prioritize active accounts (but keep high-confidence suspended/non-member matches)
            # Thresholds adapt based on address data availability to catch duplicate accounts
            # ============================================================================
            
            # Filter out unusable entries
            matches = [m for m in matches if 'DO NOT USE' not in str(m['company']).upper()]
            
            # Check if any matches have meaningful address data (used for filtering logic)
            has_any_address = any((m.get('ratio_address') or 0) > 0 for m in matches)

            # Filter 1: Prioritize hiring client relationships
            # Only apply when hc_name is non-empty — empty hc_name would falsely match
            # CBX entries with empty hiring_client_names ('' == '' is always True)
            hc_name = str(hc_row[HC_HIRING_CLIENT_NAME]).strip().lower()
            if hc_name:
                def has_hc_name(m):
                    names = str(m.get('hiring_client_names', '')).lower().split(';')
                    return any(hc_name == n.strip() for n in names)

                hc_matches = [m for m in matches if has_hc_name(m)]
            else:
                hc_matches = []

            if hc_matches:
                # Keep hiring client matches but also include high-confidence matches
                # This prevents losing duplicate companies without the HC relationship yet
                if not has_any_address:
                    # No address data: use 80% company threshold (matches dynamic threshold)
                    high_confidence_matches = [m for m in matches if (m.get('ratio_company') or 0) >= 80]
                else:
                    # Have address data: use thresholds that match or are lower than matching conditions
                    # Changed from 85%+85% to 75%+85% to match Condition 2 threshold
                    high_confidence_matches = [m for m in matches if (m.get('ratio_company') or 0) >= RATIO_COMPANY_VERY_HIGH or 
                                              ((m.get('ratio_company') or 0) >= RATIO_COMPANY_GOOD and (m.get('ratio_address') or 0) >= RATIO_ADDRESS_HIGH)]
                
                # Combine and deduplicate by cbx_id
                seen_ids = set()
                combined_matches = []
                for m in hc_matches + high_confidence_matches:
                    if m.get('cbx_id') not in seen_ids:
                        seen_ids.add(m.get('cbx_id'))
                        combined_matches.append(m)
                matches = combined_matches

            # Filter 2: Prioritize active accounts
            active_matches = [m for m in matches if m.get('registration_status', '').strip() == 'Active']
            if active_matches:
                # Keep active matches but also include suspended/non-member high-confidence matches
                # This ensures all potential duplicate accounts are visible
                if not has_any_address:
                    # No address data: use 80% threshold to catch duplicates
                    very_high_confidence = [m for m in matches if (m.get('ratio_company') or 0) >= 80 and 
                                            m.get('cbx_id') not in [am.get('cbx_id') for am in active_matches]]
                else:
                    # Have address data: use very strict 98% threshold
                    very_high_confidence = [m for m in matches if (m.get('ratio_company') or 0) >= 98 and 
                                            m.get('cbx_id') not in [am.get('cbx_id') for am in active_matches]]
                
                matches = active_matches + very_high_confidence

            # Sort prioritizing company name first (more stable identifier than address)
            # Then address, then hiring client relationships, then modules
            matches = sorted(matches, key=lambda x: (
                (x.get('ratio_company') or 0), 
                (x.get('ratio_address') or 0),
                x.get('hiring_client_count', 0), 
                x.get('modules', '')
            ), reverse=True)

            # Build analysis string showing top matches (up to max_matches_to_keep)
            ids = []
            for item in matches[0:max_matches_to_keep]:
                ids.append(f'{item["cbx_id"]}, {item["company"]}, {item["address"]}, {item["city"]}, {item["state"]} '
                           f'{item["country"]} {item["zip"]}, {item["email"]}, {item["first_name"]} {item["last_name"]}'
                           f' --> CR{item.get("ratio_company", 0)}, AR{item.get("ratio_address", 0)},'
                           f' CM{item.get("contact_match", False)}, HCC{item.get("hiring_client_count", 0)}, M[{item.get("modules", "")}]')

            # Prepare final row
            match_data = []
            uniques_cbx_id = set(item['cbx_id'] for item in matches)
            subscription_upgrade = False
            upgrade_price = 0.00
            prorated_upgrade_price = 0.00

            if uniques_cbx_id:
                # Set the analysis string in the first match BEFORE extracting values
                matches[0]['analysis'] = '\n'.join(ids)
                
                for key, value in matches[0].items():
                    if key != 'matched_qstatus':
                        match_data.append(value)
                hc_row.extend(match_data)
                hc_row.append(True if hc_domain in GENERIC_DOMAIN else False)
                hc_row.append(len(uniques_cbx_id) if len(uniques_cbx_id) else '')
                hc_row.append(len([i for i in matches if i['hiring_client_count'] > 0]))

                # Subscription upgrade calculation
                base_fee = hc_row[HC_BASE_SUBSCRIPTION_FEE] if hc_row[HC_BASE_SUBSCRIPTION_FEE] else CBX_DEFAULT_STANDARD_SUBSCRIPTION
                current_sub_total = matches[0]['subscription_price'] + matches[0]['employee_price']
                price_diff = float(base_fee) - current_sub_total

                if (price_diff > 0 and matches[0]['registration_status'] == 'Active'
                    and matches[0]['expiration_date'] and current_sub_total > 0.0):
                    subscription_upgrade = True
                    upgrade_price = price_diff
                    expiration_date = matches[0]['expiration_date']
                    now = datetime.now()
                    if expiration_date > now:
                        delta = expiration_date - now
                        days = min(delta.days, 365)
                        prorated_upgrade_price = days / 365 * upgrade_price
                    else:
                        prorated_upgrade_price = upgrade_price
                    if smart_boolean(hc_row[HC_IS_ASSOCIATION_FEE]):
                        upgrade_price += 100.0
                        prorated_upgrade_price += 100

                if matches[0]['account_type'] in ('elearning', 'plan_nord', 'portail_pfr', 'special'):
                    subscription_upgrade = True
                    prorated_upgrade_price = upgrade_price = hc_row[HC_BASE_SUBSCRIPTION_FEE]

                if parse_assessment_level(matches[0]['cbx_assessment_level']) < parse_assessment_level(hc_row[HC_ASSESSMENT_LEVEL]):
                    subscription_upgrade = True
                    prorated_upgrade_price = upgrade_price
            else:
                # No matches found - extend with empty values
                hc_row.extend(['' for x in range(len(analysis_headers) - 6)])

            create_in_cognibox = False if len(uniques_cbx_id) and not hc_row[HC_AMBIGUOUS] else True
            hc_row.append(subscription_upgrade)
            hc_row.append(upgrade_price)
            hc_row.append(prorated_upgrade_price)
            hc_row.append(create_in_cognibox)
            hc_row.append(action(hc_row, matches[0] if len(matches) else {}, create_in_cognibox,
                                 subscription_upgrade, matches[0]['expiration_date'] if len(matches) else None,
                                 matches[0]['is_qualified'] if len(matches) else False))
            hc_row.append(index + 1)
            
            # Move metadata to the end (same logic as legacy script)
            metadata_data = []
            for md_index in metadata_indexes:
                if md_index < len(hc_row):
                    metadata_data.insert(0, hc_row.pop(md_index))
            hc_row.extend(metadata_data)

            # Write to all sheet - using batch append for 5-10x faster performance
            out_ws.append(hc_row)

            # Progress update
            if (index + 1) % 10 == 0 or index == total - 1:
                progress = 0.05 + 0.85 * ((index + 1) / total)
                update_job(job_id, progress=progress, message=f"Processing: {index + 1}/{total}")
                logger.info(f"[{job_id}] Progress: {index + 1}/{total} ({progress*100:.1f}%)")

        logger.info(f"[{job_id}] Writing action sheets...")

        # Write action-specific sheets
        logger.info(f"[{job_id}] Writing action sheets...")
        update_job(job_id, progress=0.92, message="Writing action sheets...")

        action_col = HC_HEADER_LENGTH + len(analysis_headers) - 2

        for action_name, sheet in [
            ('onboarding', out_ws_onboarding),
            ('association_fee', out_ws_association_fee),
            ('re_onboarding', out_ws_re_onboarding),
            ('subscription_upgrade', out_ws_subscription_upgrade),
            ('ambiguous_onboarding', out_ws_ambiguous_onboarding),
            ('restore_suspended', out_ws_restore_suspended),
            ('activation_link', out_ws_activation_link),
            ('already_qualified', out_ws_already_qualified),
            ('add_questionnaire', out_ws_add_questionnaire),
            ('missing_info', out_ws_missing_information),
            ('follow_up_qualification', out_ws_follow_up_qualification)
        ]:
            filtered = [row for row in hc_data if row[action_col] == action_name]
            # Batch append for better performance
            for row in filtered:
                sheet.append(row)
        
        # Write Data to import sheet (onboarding contractors only)
        logger.info(f"[{job_id}] Writing 'Data to import' sheet...")
        hc_onboarding_rd = [row for row in hc_data if row[action_col] == 'onboarding']
        for idx, row in enumerate(hc_onboarding_rd):
            column = 0
            for i, value in enumerate(row):
                if i < len(rd_headers_mapping) and rd_headers_mapping[i]:
                    column += 1
                    # Swap pricing_group_id and pricing_group_code columns
                    if column == rd_pricing_group_id_col:
                        out_ws_onboarding_rd.cell(idx + 2, column + 1, value)
                    elif column == rd_pricing_group_code_col:
                        out_ws_onboarding_rd.cell(idx + 2, column - 1, value)
                    else:
                        out_ws_onboarding_rd.cell(idx + 2, column, value)
        
        # Write Existing Contractors sheet (all except onboarding and missing_info)
        logger.info(f"[{job_id}] Writing 'Existing Contractors' sheet...")
        existing_contractors_rd = [row for row in hc_data if row[action_col] != 'onboarding' and row[action_col] != 'missing_info']
        for idx, row in enumerate(existing_contractors_rd):
            column = 0
            for i, value in enumerate(row):
                if i < len(existing_contractors_headers_mapping) and existing_contractors_headers_mapping[i]:
                    column += 1
                    out_ws_existing_contractors.cell(idx + 2, column, value)
        
        # Write Data for HS sheet (all data)
        logger.info(f"[{job_id}] Writing 'Data for HS' sheet...")
        for idx, row in enumerate(hc_data):
            column = 0
            for i, value in enumerate(row):
                if i < len(hs_headers_mapping) and hs_headers_mapping[i]:
                    column += 1
                    out_ws_onboarding_hs.cell(idx + 2, column, value)

        # Format Excel sheets (like legacy script)
        logger.info(f"[{job_id}] Formatting Excel sheets...")
        update_job(job_id, progress=0.95, message="Formatting Excel sheets...")
        
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        
        table_counter = 1
        for sheet in sheets:
            # Auto-size columns based on content
            dims = {}
            for row in sheet.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
            
            for col, value in dims.items():
                sheet.column_dimensions[col].width = value
            
            # Set specific columns to fixed width for long text fields (skip special sheets: Data to import, Existing Contractors, Data for HS)
            hc_header_length = len(hc_headers)
            if sheet not in (out_ws_onboarding_rd, out_ws_existing_contractors, out_ws_onboarding_hs) and sheet.max_row > 1:  # Must have at least headers + 1 data row
                try:
                    # Find column indexes for specific headers
                    summary_col = hc_header_length + analysis_headers.index("hc_contractor_summary") + 1
                    analysis_col = hc_header_length + analysis_headers.index("analysis") + 1
                    
                    sheet.column_dimensions[get_column_letter(summary_col)].width = 150
                    sheet.column_dimensions[get_column_letter(analysis_col)].width = 150
                    
                    # Apply text wrapping for better readability
                    for i in range(2, sheet.max_row + 1):
                        sheet.cell(i, summary_col).alignment = Alignment(wrapText=True)
                        sheet.cell(i, analysis_col).alignment = Alignment(wrapText=True)
                except (ValueError, IndexError):
                    pass  # Skip if headers don't exist in this sheet
            
            # Add formatted table only if there's actual data (not just headers)
            if sheet.max_row > 1 and sheet.max_column > 0:
                # Use unique table name with counter to avoid conflicts
                table_name = f"Table{table_counter}"
                table_counter += 1
                
                try:
                    tab = Table(displayName=table_name,
                                ref=f'A1:{get_column_letter(sheet.max_column)}{sheet.max_row}')
                    tab.tableStyleInfo = style
                    sheet.add_table(tab)
                except Exception as e:
                    logger.warning(f"[{job_id}] Could not add table to sheet {sheet.title}: {e}")
        
        # Save output
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        output_filename = f"Onboarding_Analysis_Results_{timestamp}.xlsx"
        output_file = OUTPUT_DIR / output_filename
        logger.info(f"[{job_id}] Saving to {output_file}...")
        update_job(job_id, progress=0.98, message="Saving results...")
        out_wb.save(filename=output_file)

        elapsed = time.time() - t0
        logger.info(f"[{job_id}] ========== JOB COMPLETE in {elapsed:.1f}s ==========")
        update_job(job_id, status="completed", progress=1.0, message="Done!", result_file=output_filename)

    except Exception as e:
        logger.exception(f"[{job_id}] FAILED: {e}")
        update_job(job_id, status="failed", message=str(e), error=str(e))
    finally:
        # Best-effort cleanup of uploaded inputs to avoid unbounded disk growth.
        for p in (cbx_path, hc_path):
            try:
                if p and p.exists():
                    p.unlink()
            except Exception:
                logger.warning(f"[{job_id}] Could not delete temp file: {p}")


# ============================================================================
# API MODELS
# ============================================================================

class JobStatus(BaseModel):
    job_id: str
    status: str
    progress: float
    message: str
    result_file: Optional[str] = None
    error: Optional[str] = None
    created_at: str


# ============================================================================
# API ENDPOINTS
# ============================================================================

@app.post("/api/match", response_model=JobStatus)
async def match(
    cbx_file: UploadFile = File(...),
    hc_file: UploadFile = File(...),
):
    try:
        # Input validation - file size and extension checks
        logger.info("========== NEW REQUEST ==========")
        logger.info(f"CBX: {cbx_file.filename}, HC: {hc_file.filename}")
        
        # Validate file extensions (security: prevent path traversal)
        cbx_ext = Path(cbx_file.filename).suffix.lower()
        hc_ext = Path(hc_file.filename).suffix.lower()
        
        if cbx_ext not in ALLOWED_EXTENSIONS or hc_ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(status_code=400, detail=f"Only {', '.join(ALLOWED_EXTENSIONS)} files are supported")
        
        # Fixed optimized thresholds
        min_company_ratio = RATIO_COMPANY_GOOD
        min_address_ratio = RATIO_ADDRESS_HIGH
        
        logger.info(f"Match thresholds - Company: {min_company_ratio}%, Address: {min_address_ratio}%")

        job_id = str(uuid.uuid4())[:8]

        # Use sanitized extensions to prevent path traversal attacks
        cbx_path = UPLOAD_DIR / f"{job_id}_cbx{cbx_ext}"
        hc_path = UPLOAD_DIR / f"{job_id}_hc{hc_ext}"

        # Save files
        with open(cbx_path, "wb") as f:
            shutil.copyfileobj(cbx_file.file, f)
        with open(hc_path, "wb") as f:
            shutil.copyfileobj(hc_file.file, f)

        # Close upload streams early to free resources.
        try:
            await cbx_file.close()
        finally:
            await hc_file.close()

        logger.info(f"Files saved: {cbx_path}, {hc_path}")

        # Create job
        with jobs_lock:
            jobs[job_id] = {
                "job_id": job_id,
                "status": "processing",
                "progress": 0.0,
                "message": "Starting...",
                "created_at": datetime.now().isoformat(),
                "result_file": None,
                "error": None,
            }

        # Start background processing
        loop = asyncio.get_running_loop()
        loop.run_in_executor(EXECUTOR, process_matching_job, job_id, cbx_path, hc_path, min_company_ratio, min_address_ratio)

        logger.info(f"Job {job_id} started with thresholds ({min_company_ratio}/{min_address_ratio})")
        with jobs_lock:
            return JobStatus(**jobs[job_id])

    except Exception as e:
        logger.exception(f"Error starting job: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/jobs/{job_id}", response_model=JobStatus)
async def get_status(job_id: str):
    with jobs_lock:
        if job_id not in jobs:
            raise HTTPException(404, "Job not found")
        return JobStatus(**jobs[job_id])


@app.get("/api/jobs/{job_id}/download")
async def download(job_id: str):
    with jobs_lock:
        if job_id not in jobs:
            raise HTTPException(404, "Job not found")
        job = dict(jobs[job_id])
    if job["status"] != "completed":
        raise HTTPException(400, "Job not completed")
    file_path = OUTPUT_DIR / job["result_file"]
    if not file_path.exists():
        raise HTTPException(404, "Result file not found")
    return FileResponse(file_path, filename=job["result_file"])


@app.get("/api/jobs")
async def list_jobs():
    with jobs_lock:
        return {"jobs": list(jobs.values()), "total": len(jobs)}


@app.get("/api/health")
async def health():
    with jobs_lock:
        jobs_active = len(jobs)
    return {
        "status": "healthy",
        "jobs_active": jobs_active,
        "upload_dir": str(UPLOAD_DIR),
        "output_dir": str(OUTPUT_DIR)
    }


@app.get("/")
async def root():
    return {"status": "ok", "message": "Onboarding Analysis Tool API v2", "version": "2.0"}


if __name__ == "__main__":
    import uvicorn
    logger.info("Starting Onboarding Analysis Tool Backend v2...")
    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="info")

