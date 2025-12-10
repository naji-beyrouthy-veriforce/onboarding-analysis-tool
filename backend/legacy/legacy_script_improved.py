import argparse
import csv
import re
import string
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
try:
    from rapidfuzz import fuzz
except ImportError:
    from fuzzywuzzy import fuzz
from datetime import datetime, timedelta
from convertTimeZone import convertFromIANATimezone

CBX_DEFAULT_STANDARD_SUBSCRIPTION = 803
CBX_HEADER_LENGTH = 28
# noinspection SpellCheckingInspection
CBX_ID, CBX_COMPANY_FR, CBX_COMPANY_EN, CBX_COMPANY_OLD, CBX_ADDRESS, CBX_CITY, CBX_STATE, \
    CBX_COUNTRY, CBX_ZIP, CBX_FISTNAME, CBX_LASTNAME, CBX_EMAIL, CBX_EXPIRATION_DATE, CBX_REGISTRATION_STATUS, \
    CBX_SUSPENDED, CBX_MODULES, CBX_ACCESS_MODES, CBX_ACCOUNT_TYPE, CBX_SUB_PRICE_CAD, CBX_EMPL_PRICE_CAD,\
    CBX_SUB_PRICE_USD, CBX_EMPL_PRICE_USD, CBX_HIRING_CLIENT_NAMES, \
    CBX_HIRING_CLIENT_IDS, CBX_HIRING_CLIENT_QSTATUS, CBX_PARENTS, CBX_ASSESSMENT_LEVEL, CBX_IS_NEW_PRODUCT = range(CBX_HEADER_LENGTH)

HC_HEADER_LENGTH = 41
HC_COMPANY, HC_FIRSTNAME, HC_LASTNAME, HC_EMAIL, HC_CONTACT_PHONE, HC_CONTACT_LANGUAGE, HC_STREET, HC_CITY, \
    HC_STATE, HC_COUNTRY, HC_ZIP, HC_CATEGORY, HC_DESCRIPTION, HC_PHONE, HC_EXTENSION, HC_FAX,  HC_WEBSITE,\
    HC_LANGUAGE, HC_IS_TAKE_OVER, HC_TAKEOVER_QUALIFICATION_DATE, HC_TAKEOVER_QF_STATUS, \
    HC_PROJECT_NAME, HC_QUESTIONNAIRE_NAME, HC_QUESTIONNAIRE_ID, HC_PRICING_GROUP_ID, HC_PRICING_GROUP_CODE, \
    HC_HIRING_CLIENT_NAME, HC_HIRING_CLIENT_ID, HC_IS_ASSOCIATION_FEE, HC_BASE_SUBSCRIPTION_FEE, \
    HC_CONTACT_CURRENCY, HC_AGENT_IN_CHARGE_ID, HC_TAKEOVER_FOLLOW_UP_DATE, HC_TAKEOVER_RENEWAL_DATE, \
    HC_INFORMATION_SHARED, HC_CONTACT_TIMEZONE, HC_DO_NOT_MATCH, HC_FORCE_CBX_ID, HC_AMBIGUOUS, \
    HC_CONTRACTORCHECK_ACCOUNT, HC_ASSESSMENT_LEVEL \
    = range(HC_HEADER_LENGTH)

SUPPORTED_CURRENCIES = ('CAD', 'USD')

assessment_levels = {
    "gold": 2,
    "silver": 2,
    "bronze" : 1,
    "level3": 2, 
    "level2": 2,
    "level1": 1,
    "3":2,
    "2":2,
    "1":1
}

# Used in order to switch code and id in data to import
rd_pricing_group_id_col = -1
rd_pricing_group_code_col = -1

# noinspection SpellCheckingInspection
cbx_headers = ['id', 'name_fr', 'name_en', 'old_names', 'address', 'city', 'state', 'country', 'postal_code',
               'first_name', 'last_name', 'email', 'cbx_expiration_date', 'registration_code', 'suspended',
               'modules', 'access_modes', 'code', 'subscription_price_cad', 'employee_price_cad',
               'subscription_price_usd', 'employee_price_usd', 'hiring_client_names',
               'hiring_client_ids', 'hiring_client_qstatus', 'parents', 'assessment_level', 'new_product']

# noinspection SpellCheckingInspection
hiring_client_headers = ['contractor_name', 'contact_first_name', 'contact_last_name', 'contact_email', 'contact_phone',
              'contact_language', 'address', 'city', 'province_state_iso2', 'country_iso2',
              'postal_code', 'category', 'description', 'phone', 'extension', 'fax', 'website', 'language',
              'is_take_over', 'qualification_expiration_date',
              'qualification_status', 'batch', 'questionnaire_name', 'questionnaire_id',
              'pricing_group_id', 'pricing_group_code', 'hiring_client_name', 'hiring_client_id', 'is_association_fee',
              'base_subscription_fee', 'contact_currency', 'agent_in_charge_id', 'take_over_follow-up_date',
              'renewal_date', 'information_shared', 'contact_timezone', 'do_not_match',
              'force_cbx_id', 'ambiguous', 'contractorcheck_account', 'assessment_level']

# noinspection SpellCheckingInspection
analysis_headers = ['cbx_id', 'hc_contractor_summary', 'analysis','cbx_contractor', 'cbx_street', 'cbx_city', 'cbx_state', 'cbx_zip', 'cbx_country',
                    'cbx_expiration_date', 'registration_status', 'suspended', 'cbx_email',
                    'cbx_first_name', 'cbx_last_name', 'modules', 'cbx_account_type',
                    'cbx_subscription_fee', 'cbx_employee_price', 'parents', 'previous',
                    'hiring_client_names', 'hiring_client_count',
                    'is_in_relationship', 'is_qualified', 'ratio_company', 'ratio_address',
                    'contact_match', 'cbx_assessment_level', 'new_product', 'generic_domain', 'match_count', 'match_count_with_hc',
                    'is_subscription_upgrade', 'upgrade_price', 'prorated_upgrade_price', 'create_in_cbx',
                    'action', 'index']

rd_headers = ['contractor_name', 'contact_first_name', 'contact_last_name', 'contact_email', 'contact_phone',
              'contact_language', 'address', 'city', 'province_state_iso2', 'country_iso2',
              'postal_code', 'description', 'phone', 'extension', 'fax', 'website', 'language',
              'qualification_expiration_date', 'qualification_status', 'contact_currency',
              'agent_in_charge_id', 'renewal_date', 'information_shared', 'contact_timezone', 'questionnaire_name', 'questionnaire_ids',
              'pricing_group_code', 'pricing_group_id', 'hiring_client_id', 'contractorcheck_account', 'assessment_level']

existing_contractors_headers = ['cbx_id']
existing_contractors_headers.extend(rd_headers.copy())

hubspot_headers = ['contractor_name', 'contact_first_name', 'contact_last_name', 'contact_email', 'contact_phone',
              'contact_language', 'address', 'city', 'province_state_iso2', 'country_iso2',
              'postal_code', 'cbx_id', 'cbx_expiration_date', 'questionnaire_name',
              'questionnaire_id', 'hiring_client_name', 'hiring_client_id', 'action']


metadata_headers = ['metadata_x', 'metadata_y', 'metadata_z', '...']

# noinspection SpellCheckingInspection
BASE_GENERIC_DOMAIN = ['yahoo.ca', 'yahoo.com', 'hotmail.com', 'gmail.com', 'outlook.com',
                       'bell.com', 'bell.ca', 'videotron.ca', 'eastlink.ca', 'kos.net', 'bellnet.ca', 'sasktel.net',
                       'aol.com', 'tlb.sympatico.ca', 'sogetel.net', 'cgocable.ca',
                       'hotmail.ca', 'live.ca', 'icloud.com', 'hotmail.fr', 'yahoo.com', 'outlook.fr', 'msn.com',
                       'globetrotter.net', 'live.com', 'sympatico.ca', 'live.fr', 'yahoo.fr', 'telus.net',
                       'shaw.ca', 'me.com', 'bell.net', 'cablevision.qc.ca', 'live.ca', 'tlb.sympatico.ca',
                       '', 'videotron.qc.ca', 'ivic.qc.ca', 'qc.aira.com', 'canada.ca', 'axion.ca', 'bellsouth.net', 
                       'telusplanet.net','rogers.com', 'mymts.net', 'nb.aibn.com', 'on.aibn.com', 'live.be', 'nbnet.nb.ca',
                       'execulink.com', 'bellaliant.com', 'nf.aibn.com', 'clintar.com', 'pathcom.com', 'oricom.ca', 'mts.net',
                       'xplornet.com', 'mcsnet.ca', 'att.net', 'ymail.com', 'mail.com', 'bellaliant.net', 'ns.sympatico.ca', 
                       'ns.aliantzinc.ca', 'mnsi.net']
# noinspection SpellCheckingInspection
BASE_GENERIC_COMPANY_NAME_WORDS = ['construction', 'contracting', 'industriel', 'industriels', 'service',
                                   'services', 'inc', 'limited', 'ltd', 'ltee', 'ltée', 'co', 'industrial',
                                   'solutions', 'llc', 'enterprises', 'systems', 'industries',
                                   'technologies', 'company', 'corporation', 'installations', 'enr']

# Compiled regex pattern for remove_generics() optimization
GENERIC_WORDS_PATTERN = None  # Will be compiled after args are parsed


def norm_name(name):
    """Normalize hiring client name by removing punctuation and converting to lowercase."""
    if not name:
        return ''
    return str(name).translate(str.maketrans('', '', string.punctuation)).strip().lower()


def chunks(lst, n):
    """Yield successive n-sized chunks from lst."""
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


hiring_client_headers_with_metadata = hiring_client_headers.copy()
hiring_client_headers_with_metadata.extend(metadata_headers)
cbx_headers_text = '\n'.join([', '.join(x) for x in list(chunks(cbx_headers, 5))])
hiring_client_headers_text = '\n'.join([', '.join(x) for x in list(chunks(hiring_client_headers_with_metadata, 5))])
analysis_headers_text = '\n'.join([', '.join(x) for x in list(chunks(analysis_headers, 5))])
existing_contractors_text = '\n'.join([', '.join(x) for x in list(chunks(existing_contractors_headers, 5))])

if len(hiring_client_headers) != HC_HEADER_LENGTH:
    raise AssertionError('hc header inconsistencies')

if len(cbx_headers) != CBX_HEADER_LENGTH:
    raise AssertionError('cbx header inconsistencies')

# define commandline parser
parser = argparse.ArgumentParser(
    description='Tool to match contractor list provided by hiring clients to business units in CBX, '
                'all input/output files must be in the current directory',
    formatter_class=argparse.RawTextHelpFormatter)
parser.add_argument('cbx_list',
                    help=f'csv DB export file of business units with the following columns:\n{cbx_headers_text}\n\n')

parser.add_argument('hc_list',
                    help=f'xlsx file of the hiring client contractors and the '
                         f'following columns:\n{hiring_client_headers_text}\n\n')
parser.add_argument('output',
                    help=f'the xlsx file to be created with the hc_list columns and the following analysis columns:'
                         f'\n{analysis_headers_text}\n\n**Please note that metadata columns from the'
                         f' hc file are moved after the analysis data')


parser.add_argument('--hc_list_sheet_name', dest='hc_list_sheet_name', action='store',
                    default=None,
                    help='specify the sheet in the excel file where the hiring client data is located'
                         ' (default separator is the active sheet)')

parser.add_argument('--hc_list_offset', dest='hc_list_offset', action='store',
                    default=None,
                    help='offset where the data starts in the form of <row>,<column> (default is 1,1). '
                         'This includes the headers')

parser.add_argument('--min_company_match_ratio', dest='ratio_company', action='store',
                    default=80,
                    help='Minimum match ratio for contractors, between 0 and 100 (default 80)')

parser.add_argument('--min_address_match_ratio', dest='ratio_address', action='store',
                    default=80,
                    help='Minimum match ratio for addresses (street + zip), between 0 and 100 (default 80)')

parser.add_argument('--additional_generic_domain', dest='additional_generic_domain', action='store',
                    default='',
                    help='list of domains to ignore separated by the list separator (default separator is ;)')

parser.add_argument('--additional_generic_name_word', dest='additional_generic_name_word', action='store',
                    default='',
                    help='list of generic words in company name to ignore separated by the list separator'
                         ' (default separator is ;)')

parser.add_argument('--cbx_list_encoding', dest='cbx_encoding', action='store',
                    default='utf-8-sig',
                    help='Encoding for the cbx list (default: utf-8-sig)')

parser.add_argument('--list_separator', dest='list_separator', action='store',
                    default=';',
                    help='string separator used for lists (default: ;)')

parser.add_argument('--no_headers', dest='no_headers', action='store_true',
                    help='to indicate that input files have no headers')

parser.add_argument('--ignore_warnings', dest='ignore_warnings', action='store_true',
                    help='to ignore data consistency checks and run anyway...')

args = parser.parse_args()
GENERIC_DOMAIN = BASE_GENERIC_DOMAIN + args.additional_generic_domain.split(args.list_separator)
GENERIC_COMPANY_NAME_WORDS = BASE_GENERIC_COMPANY_NAME_WORDS + \
                             args.additional_generic_name_word.split(args.list_separator)

# Compile regex pattern for optimal performance in remove_generics()
GENERIC_WORDS_PATTERN = re.compile(
    r'\b(?:' + '|'.join(re.escape(word) for word in GENERIC_COMPANY_NAME_WORDS) + r')\b',
    re.IGNORECASE
)


def smart_boolean(bool_data):
    if isinstance(bool_data, str):
        bool_data = bool_data.lower().strip()
        return True if bool_data in ('true', '=true', 'yes', 'vraie', '=vraie', '1') else False
    else:
        return bool(bool_data)


# noinspection PyShadowingNames
def add_analysis_data(hc_row, cbx_row, hc_email, ratio_company=None, ratio_address=None, contact_match=None):
    cbx_company = cbx_row[CBX_COMPANY_FR] or cbx_row[CBX_COMPANY_EN]
    # Removed print for production - use logging if needed
    # print('   --> ', cbx_company, hc_email, cbx_row[CBX_ID], ratio_company, ratio_address, contact_match)

    hiring_clients_list = [norm_name(x) for x in cbx_row[CBX_HIRING_CLIENT_NAMES].split(args.list_separator)]
    hiring_clients_qstatus = cbx_row[CBX_HIRING_CLIENT_QSTATUS].split(args.list_separator)
    hc_name_norm = norm_name(hc_row[HC_HIRING_CLIENT_NAME])
    hc_count = len(hiring_clients_list) if cbx_row[CBX_HIRING_CLIENT_NAMES] else 0
    is_in_relationship = hc_name_norm in hiring_clients_list and hc_name_norm != ''
    is_qualified = False
    matched_qstatus = None
    
    # Cache currency check and optimize price conversions
    is_cad = hc_row[HC_CONTACT_CURRENCY] == "CAD"
    sub_price_usd = float(cbx_row[CBX_SUB_PRICE_USD]) if cbx_row[CBX_SUB_PRICE_USD] else 0.0
    employee_price_usd = float(cbx_row[CBX_EMPL_PRICE_USD]) if cbx_row[CBX_EMPL_PRICE_USD] else 0.0
    sub_price_cad = float(cbx_row[CBX_SUB_PRICE_CAD]) if cbx_row[CBX_SUB_PRICE_CAD] else 0.0
    employee_price_cad = float(cbx_row[CBX_EMPL_PRICE_CAD]) if cbx_row[CBX_EMPL_PRICE_CAD] else 0.0
    subscription_price = sub_price_cad if is_cad else sub_price_usd
    employee_price = employee_price_cad if is_cad else employee_price_usd
    hiring_client_contractor_summary = f'{hc_row[HC_COMPANY]}, {hc_row[HC_STREET]}, {hc_row[HC_CITY]}, {hc_row[HC_STATE]}, {hc_row[HC_COUNTRY]}, {hc_row[HC_ZIP]}, {hc_row[HC_EMAIL]}, {hc_row[HC_FIRSTNAME]} {hc_row[HC_LASTNAME]}'

    if hc_row[HC_CONTACT_CURRENCY] != '' and hc_row[HC_CONTACT_CURRENCY] not in SUPPORTED_CURRENCIES:
        raise AssertionError(f'Invalid currency: {hc_row[HC_CONTACT_CURRENCY]}, must be in {SUPPORTED_CURRENCIES}')
    for idx, val in enumerate(hiring_clients_list):
        if val == hc_name_norm and idx < len(hiring_clients_qstatus):
            qstatus = hiring_clients_qstatus[idx].strip().lower()
            matched_qstatus = qstatus
            if qstatus == 'validated':
                is_qualified = True
            break
    try:
        expiration_date = datetime.strptime(cbx_row[CBX_EXPIRATION_DATE],
                                        "%d/%m/%y") if cbx_row[CBX_EXPIRATION_DATE] else None
    except ValueError:
        expiration_date = datetime.strptime(cbx_row[CBX_EXPIRATION_DATE],
                                        "%d/%m/%Y") if cbx_row[CBX_EXPIRATION_DATE] else None

    return {'cbx_id': int(cbx_row[CBX_ID]), 'hc_contractor_summary': hiring_client_contractor_summary, 'analysis':'', 'company': cbx_company, 'address': cbx_row[CBX_ADDRESS],
        'city': cbx_row[CBX_CITY], 'state': cbx_row[CBX_STATE], 'zip': cbx_row[CBX_ZIP],
        'country': cbx_row[CBX_COUNTRY], 'expiration_date': expiration_date,
        'registration_status': cbx_row[CBX_REGISTRATION_STATUS],
        'suspended': cbx_row[CBX_SUSPENDED], 'email': cbx_row[CBX_EMAIL], 'first_name': cbx_row[CBX_FISTNAME],
        'last_name': cbx_row[CBX_LASTNAME], 'modules': cbx_row[CBX_MODULES],
        'account_type': cbx_row[CBX_ACCOUNT_TYPE],
        'subscription_price': subscription_price,
        'employee_price': employee_price,
        'parents': cbx_row[CBX_PARENTS], 'previous': cbx_row[CBX_COMPANY_OLD],
        'hiring_client_names': cbx_row[CBX_HIRING_CLIENT_NAMES], 'hiring_client_count': hc_count,
        'is_in_relationship': is_in_relationship, 'is_qualified': is_qualified,
        'matched_qstatus': matched_qstatus,
        'ratio_company': ratio_company, 'ratio_address': ratio_address, 'contact_match': contact_match, 
        'cbx_assessment_level': cbx_row[CBX_ASSESSMENT_LEVEL],
        'new_product': cbx_row[CBX_IS_NEW_PRODUCT]
        }


def core_mandatory_provided(hcd):
    mandatory_fields = (HC_COMPANY, HC_FIRSTNAME, HC_LASTNAME, HC_EMAIL, HC_CONTACT_PHONE,
                        HC_STREET, HC_CITY, HC_STATE, HC_COUNTRY, HC_ZIP)
    country = hcd[HC_COUNTRY].strip().lower() if isinstance(hcd[HC_COUNTRY], str) else hcd[HC_COUNTRY]
    for field in mandatory_fields:
        f_value = hcd[field].strip() if isinstance(hcd[field], str) else hcd[field]
        if f_value == "":
            if field == HC_STATE and country not in ('ca', 'us'):
                pass
            else:
                return False
    return True


# noinspection PyShadowingNames
def action(hc_data, cbx_data, create, subscription_update, expiration_date, is_qualified, ignore):
    if create:
        if smart_boolean(hc_data[HC_IS_TAKE_OVER]):
            return 'activation_link'
        else:
            if hc_data[HC_AMBIGUOUS]:
                return 'ambiguous_onboarding'
            elif core_mandatory_provided(hc_data):
                return 'onboarding'
            else:
                return 'missing_info'
    else:
        reg_status = cbx_data['registration_status']
        if smart_boolean(hc_data[HC_IS_TAKE_OVER]):
            if reg_status == 'Suspended':
                return 'restore_suspended'
            elif reg_status == 'Active':
                return 'add_questionnaire'
            elif reg_status == 'Non Member':
                return 'activation_link'
            else:
                print(f'WARNING: invalid registration status {reg_status}')
                if not ignore:
                    raise ValueError(f'Invalid registration status: {reg_status}')
        else:
            if reg_status == 'Active':
                if cbx_data['is_in_relationship']:
                    qstatus = cbx_data.get('matched_qstatus', None)
                    if qstatus == 'validated':
                        return 'already_qualified'
                    elif qstatus in ('pending', 'expired', 'conditional', 'refused'):
                        return 'follow_up_qualification'
                    else:
                        # If qstatus is missing or unknown, fallback to previous logic
                        if is_qualified:
                            return 'already_qualified'
                        else:
                            return 'follow_up_qualification'
                else:
                    if subscription_update:
                        return 'subscription_upgrade'
                    elif hc_data[HC_IS_ASSOCIATION_FEE] and not cbx_data['is_in_relationship']:
                        # Association fee only if renewal > 60 days, else add questionnaire
                        if expiration_date:
                            in_60_days = datetime.now() + timedelta(days=60)
                            if expiration_date > in_60_days:
                                return 'association_fee'
                            else:
                                return 'add_questionnaire'
                        else:
                            return 'association_fee'
                    else:
                        return 'add_questionnaire'
            elif reg_status == 'Suspended':
                return 'restore_suspended'
            elif reg_status in ('Non Member', '', None):
                return 're_onboarding'
            else:
                raise AssertionError(f'invalid registration status: {reg_status}')


def remove_generics(company_name):
    """Remove generic company name words using pre-compiled regex pattern."""
    if GENERIC_WORDS_PATTERN:
        return GENERIC_WORDS_PATTERN.sub('', company_name)
    # Fallback if pattern not compiled yet
    for word in GENERIC_COMPANY_NAME_WORDS:
        company_name = re.sub(r'\b' + word + r'\b', '', company_name)
    return company_name


# noinspection PyShadowingNames
def check_headers(headers, standards, ignore):
    headers = [x.lower().strip() for x in headers]
    for idx, val in enumerate(standards):
        if val != headers[idx]:
            print(f'WARNING: got "{headers[idx]}" while expecting "{val}" in column {idx + 1}')
            if not ignore:
                exit(-1)


def clean_company_name(name):
    name = name.lower().replace('.', '').replace(',', '').strip()
    name = re.sub(r"\([^()]*\)", "", name)
    name = remove_generics(name)
    return name


def parse_assessment_level(level):
    """Parse assessment level from various input formats."""
    if level is None or (isinstance(level, int) and 0 < level < 4):
        return level
    
    if isinstance(level, str):
        level_key = level.lower()
        return assessment_levels.get(level_key, 0)
    
    return 0

if __name__ == '__main__':
    data_path = './data/'
    cbx_file = data_path + args.cbx_list
    hc_file = data_path + args.hc_list
    output_file = data_path + args.output

    # output parameters used
    print(f'Starting at {datetime.now()}')
    print(f'Reading CBX list: {args.cbx_list} [{args.cbx_encoding}]')
    print(f'Reading HC list: {args.hc_list}')
    print(f'Outputting results in: {args.output}')
    print(f'contractor match ratio: {args.ratio_company}')
    print(f'address match ratio: {args.ratio_address}')
    print(f'list of generic domains:\n{BASE_GENERIC_DOMAIN}')
    print(f'additional generic domain: {args.additional_generic_domain}')
    # read data
    cbx_data = []
    hc_data = []
    hc_row = []
    print('Reading Cognibox data file...')
    with open(cbx_file, 'r', encoding=args.cbx_encoding) as cbx:
        for row in csv.reader(cbx):
            cbx_data.append(row)
    # check cbx db ata consistency
    if cbx_data and len(cbx_data[0]) != len(cbx_headers):
        print(f'WARNING: got {len(cbx_data[0])} columns when expecting {len(cbx_headers)}')
        if not args.ignore_warnings:
            exit(-1)
    if not args.no_headers:
        headers = cbx_data.pop(0)
        headers = [x.lower().strip() for x in headers]
        check_headers(headers, cbx_headers, args.ignore_warnings)
    # for index, row in enumerate(cbx_data):
    #     access_modes = row[CBX_ACCESS_MODES].split(';')
    #     # only keep contractors on Non-member without any access mode (ignore training and hiring clients)
    #     if 'Contractor' not in access_modes and access_modes:
    #         cbx_data.pop(index)
    print(f'Completed reading {len(cbx_data)} contractors.')
    
    # Pre-normalize CBX data once to avoid millions of repeated computations
    print('Pre-normalizing CBX data for faster matching...')
    for cbx_row in cbx_data:
        cbx_row.append(clean_company_name(cbx_row[CBX_COMPANY_EN]))  # Index 28
        cbx_row.append(clean_company_name(cbx_row[CBX_COMPANY_FR]))  # Index 29
        cbx_row.append(cbx_row[CBX_ADDRESS].lower().replace('.', '').strip())  # Index 30
    print(f'Pre-normalization complete.')

    print('Reading hiring client data file...')
    hc_wb = openpyxl.load_workbook(hc_file, read_only=True, data_only=True)
    if args.hc_list_sheet_name:
        hc_sheet = hc_wb.get_sheet_by_name(args.hc_list_sheet_name)
    else:
        hc_sheet = hc_wb.active
    max_row = hc_sheet.max_row
    max_column = hc_sheet.max_column
    row_offset = 0 if not args.hc_list_offset else int(args.hc_list_offset.split(',')[0])-1
    column_offset = 0 if not args.hc_list_offset else int(args.hc_list_offset.split(',')[1])-1

    if max_column > 250 or max_row > 10000:
        print(f'WARNING: File is large: {max_row} rows and {max_column}. must be less than 10000 and 250')
        if not args.ignore_warnings:
            exit(-1)
    for row in hc_sheet.rows:
        # start data retrieval at offset
        while row_offset:
            next(hc_sheet.rows)
            row_offset -= 1
        row = row[column_offset:]
        # retrieve
        if not row[0].value:
            continue
        hc_data.append([cell.value if cell.value is not None else '' for cell in row])
    total = len(hc_data) - 1
    metadata_indexes = []
    headers = []
    rd_headers_mapping = []
    hs_headers_mapping = []
    existing_contractors_headers_mapping = []
    # check hc data consistency
    if hc_data and len(hc_data[0]) < len(hiring_client_headers):
        print(f'WARNING: got {len(hc_data[0])} columns when at least {len(hiring_client_headers)} is expected')
        if not args.ignore_warnings:
            exit(-1)
    if not args.no_headers:
        headers = hc_data.pop(0)
        headers = [x.lower().strip() for x in headers]
        check_headers(headers, hiring_client_headers, args.ignore_warnings)
    else:
        if hc_data and len(hc_data[0]) != len(hiring_client_headers):
            print(f'WARNING: got {len(hc_data[0])} columns when {len(hiring_client_headers)} is exactly expected')
            if not args.ignore_warnings:
                exit(-1)
    # checking currency integrity and strip characters from contact phone
    for row in hc_data:
        if row[HC_COUNTRY].lower().strip() == 'ca':
            if row[HC_CONTACT_CURRENCY].lower().strip() not in ('cad', ''):
                print(f'WARNING: currency and country mismatch: {row[HC_CONTACT_CURRENCY]} and'
                      f' "{row[HC_COUNTRY]}". Expected CAD in row {row}')
                if not args.ignore_warnings:
                    exit(-1)
        elif row[HC_COUNTRY].lower().strip() != '':
            if row[HC_CONTACT_CURRENCY].lower().strip() not in ('usd', ''):
                print(f'WARNING: currency and country mismatch: {row[HC_CONTACT_CURRENCY]} and'
                      f' "{row[HC_COUNTRY]}". Expected USD in row {row}')
                if not args.ignore_warnings:
                    exit(-1)
        row[HC_EMAIL] = str(row[HC_EMAIL]).strip()
        # correct and normalize phone number
        extension = ''
        if isinstance(row[HC_CONTACT_PHONE], str):
            for x in ('ext', 'x', 'poste', ',', 'p'):
                f_index = row[HC_CONTACT_PHONE].lower().find(x)
                if f_index >= 0:
                    extension = row[HC_CONTACT_PHONE][f_index + len(x):]
                    row[HC_CONTACT_PHONE] = row[HC_CONTACT_PHONE][0:f_index]
                    break
            row[HC_CONTACT_PHONE] = re.sub("[^0-9]", "", row[HC_CONTACT_PHONE])
        elif isinstance(row[HC_CONTACT_PHONE], int):
            row[HC_CONTACT_PHONE] = str(row[HC_CONTACT_PHONE])
        if row[HC_CONTACT_PHONE] and not row[HC_PHONE]:
            row[HC_PHONE] = row[HC_CONTACT_PHONE]
            row[HC_EXTENSION] = extension
        if isinstance(row[HC_EXTENSION], str):
            row[HC_EXTENSION] = re.sub("[^0-9]", "", row[HC_EXTENSION])
        # make language lower case; currency, state ISO2 and country ISO2 upper case
        row[HC_LANGUAGE] = row[HC_LANGUAGE].lower()
        row[HC_CONTACT_LANGUAGE] = row[HC_CONTACT_LANGUAGE].lower()
        row[HC_COUNTRY] = row[HC_COUNTRY].upper()
        row[HC_STATE] = row[HC_STATE].upper()
        row[HC_CONTACT_CURRENCY] = row[HC_CONTACT_CURRENCY].upper()
        # convert date-time to windows format
        row[HC_CONTACT_TIMEZONE] = convertFromIANATimezone(row[HC_CONTACT_TIMEZONE])
    print(f'Completed reading {len(hc_data)} contractors.')
    print(f'Starting data analysis...')

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = 'all'
    out_ws_onboarding = out_wb.create_sheet(title="onboarding")
    out_ws_association_fee = out_wb.create_sheet(title="association_fee")
    out_ws_re_onboarding = out_wb.create_sheet(title="re_onboarding")
    out_ws_subscription_upgrade = out_wb.create_sheet(title="subscription_upgrade")
    out_ws_ambiguous_onboarding = out_wb.create_sheet(title="ambiguous_onboarding")
    out_ws_restore_suspended = out_wb.create_sheet(title="restore_suspended")
    out_ws_activation_link = out_wb.create_sheet(title="activation_link")
    out_ws_already_qualified = out_wb.create_sheet(title="already_qualified")
    out_ws_add_questionnaire = out_wb.create_sheet(title="add_questionnaire")
    out_ws_missing_information = out_wb.create_sheet(title="missing_info")
    out_ws_follow_up_qualification = out_wb.create_sheet(title="follow_up_qualification")
    out_ws_onboarding_rd = out_wb.create_sheet(title="Data to import")
    out_ws_existing_contractors = out_wb.create_sheet(title="Existing Contractors")
    out_ws_onboarding_hs = out_wb.create_sheet(title="Data for HS")

    sheets = (out_ws, out_ws_onboarding, out_ws_association_fee, out_ws_re_onboarding, out_ws_subscription_upgrade,
              out_ws_ambiguous_onboarding, out_ws_restore_suspended, out_ws_activation_link, out_ws_already_qualified,
              out_ws_add_questionnaire, out_ws_missing_information, out_ws_follow_up_qualification,
              out_ws_onboarding_rd, out_ws_existing_contractors, out_ws_onboarding_hs)

    # append analysis headers and move metadata headers at the end
    if not args.no_headers:
        for idx, val in enumerate(headers):
            if val.lower().startswith('metadata'):
                metadata_indexes.append(idx)
        metadata_indexes.sort(reverse=True)
        headers.extend(analysis_headers)
        metadata_array = []
        for md_index in metadata_indexes:
            metadata_array.insert(0, headers.pop(md_index))
        headers.extend(metadata_array)
        hubspot_headers.extend(metadata_array)  # hubspot headers must includes metadata if present
        existing_contractors_headers.extend(metadata_array)  # existing contractors headers must includes metadata if present
        rd_headers.extend(metadata_array)
        column_rd = column_hs = column_existing_contractors = 0
        for index, value in enumerate(headers):
            # skip the last two sheets since they have special mapping handled below
            for sheet in sheets[:-3]:
                sheet.cell(1, index+1, value)
            rd_headers_for_value = [s for s in rd_headers if value in s]
            if rd_headers_for_value:
                column_rd += 1
                rd_headers_mapping.append(True)
                # Invert code and id columns
                if value == "pricing_group_id":
                    adjustement = 1
                    rd_pricing_group_id_col = column_rd
                elif value == "pricing_group_code":
                    adjustement = -1
                    rd_pricing_group_code_col = column_rd
                else:
                    adjustement = 0

                if value in rd_headers:
                    out_ws_onboarding_rd.cell(1, column_rd + adjustement, value)
                else:
                    out_ws_onboarding_rd.cell(1, column_rd, rd_headers_for_value[0])
            else:
                rd_headers_mapping.append(False)

            if value in hubspot_headers:
                column_hs += 1
                hs_headers_mapping.append(True)
                out_ws_onboarding_hs.cell(1, column_hs, value)
            else:
                hs_headers_mapping.append(False)

            existing_contractors_headers_for_value = [s for s in existing_contractors_headers if value in s]
            if existing_contractors_headers_for_value:
                column_existing_contractors += 1
                existing_contractors_headers_mapping.append(True)
                if value in existing_contractors_headers:
                    out_ws_existing_contractors.cell(1, column_existing_contractors, value)
                else:
                    out_ws_existing_contractors.cell(1, column_existing_contractors, existing_contractors_headers_for_value[0])
            else:
                existing_contractors_headers_mapping.append(False)
                
        out_wb.save(filename=output_file)
    # match
    for index, hc_row in enumerate(hc_data):
        matches = []
        hc_company = hc_row[HC_COMPANY]

        clean_hc_company = clean_company_name(hc_company)
        hc_email = str(hc_row[HC_EMAIL]).lower()
        # if multiple values use the first one...
        hc_email = hc_email.split(';')[0]
        hc_email = hc_email.split('\n')[0]
        hc_email = hc_email.split(',')[0]
        hc_email = hc_email.strip()
        hc_domain = hc_email[hc_email.find('@') + 1:]
        hc_zip = str(hc_row[HC_ZIP]).replace(' ', '').upper()
        hc_address = str(hc_row[HC_STREET]).lower().replace('.', '').strip()
        hc_force_cbx = str(hc_row[HC_FORCE_CBX_ID])
        if not smart_boolean(hc_row[HC_DO_NOT_MATCH]):
            if hc_force_cbx:
                cbx_row = next(filter(lambda x: x[CBX_ID].strip() == hc_force_cbx, cbx_data), None)
                if cbx_row:
                    matches.append(add_analysis_data(hc_row, cbx_row, hc_email))
            else:
                for cbx_row in cbx_data:
                    cbx_email = cbx_row[CBX_EMAIL].lower()
                    cbx_domain = cbx_email[cbx_email.find('@') + 1:]
                    contact_match = False
                    if hc_email:
                        if hc_domain in GENERIC_DOMAIN:
                            contact_match = True if cbx_email == hc_email else False
                        else:
                            contact_match = True if cbx_domain == hc_domain else False
                    else:
                        contact_match = False
                    cbx_zip = cbx_row[CBX_ZIP].replace(' ', '').upper()
                    cbx_company_en = cbx_row[28]  # Pre-cached normalized EN name
                    cbx_company_fr = cbx_row[29]  # Pre-cached normalized FR name
                    cbx_parents = cbx_row[CBX_PARENTS]
                    cbx_previous = cbx_row[CBX_COMPANY_OLD]
                    cbx_address = cbx_row[30]  # Pre-cached normalized address
                    ratio_company_fr = fuzz.token_sort_ratio(cbx_company_fr, clean_hc_company)
                    ratio_company_en = fuzz.token_sort_ratio(cbx_company_en, clean_hc_company)
                    if cbx_row[CBX_COUNTRY] != hc_row[HC_COUNTRY]:
                        ratio_zip = ratio_address = 0.0
                    else:
                        ratio_zip = fuzz.ratio(cbx_zip, hc_zip)
                        ratio_address = fuzz.token_sort_ratio(cbx_address, hc_address)
                        ratio_address = ratio_address if ratio_zip == 0 else ratio_zip if ratio_address == 0 \
                            else ratio_address * ratio_zip / 100
                    ratio_company = ratio_company_fr if ratio_company_fr > ratio_company_en else ratio_company_en
                    ratio_previous = 0
                    for item in cbx_previous.split(args.list_separator):
                        if item in (cbx_row[CBX_COMPANY_EN], cbx_row[CBX_COMPANY_FR]):
                            continue
                        item = clean_company_name(item)
                        ratio = fuzz.token_sort_ratio(item, clean_hc_company)
                        ratio_previous = ratio if ratio > ratio_previous else ratio_previous
                    ratio_company = ratio_previous if ratio_previous > ratio_company else ratio_company
                    if (contact_match or (ratio_company >= float(args.ratio_company)
                                          and ratio_address >= float(args.ratio_address))):
                        matches.append(
                            add_analysis_data(hc_row, cbx_row, hc_email, ratio_company, ratio_address, contact_match))
                    elif ratio_company >= 95.0 or (ratio_company >= float(args.ratio_company)
                                                   and ratio_address >= float(args.ratio_address)):
                        matches.append(
                            add_analysis_data(hc_row, cbx_row, hc_email, ratio_company, ratio_address, contact_match))
        ids = []
        best_match = 0
        # Exclude 'DO NOT USE' entries
        matches = [m for m in matches if 'DO NOT USE' not in m['company'].upper()]

        # Prioritize matches where the hiring client NAME is present in Cognibox hiring_client_names
        hc_name = str(hc_row[HC_HIRING_CLIENT_NAME]).strip().lower()
        def has_hc_name(m):
            names = str(m.get('hiring_client_names', '')).lower().split(';')
            return any(hc_name == n.strip() for n in names)
        hc_matches = [m for m in matches if has_hc_name(m)]
        if hc_matches:
            matches = hc_matches

        # Prefer 'Active' registration_status
        active_matches = [m for m in matches if m.get('registration_status', '').strip() == 'Active']
        if active_matches:
            matches = active_matches

        # Sort by modules and hiring_client_count, then address/company ratios
        matches = sorted(matches, key=lambda x: (x.get('modules', 0), x.get('hiring_client_count', 0), x["ratio_address"], x["ratio_company"]), reverse=True)
        for item in matches[0:10]:
            ids.append(f'{item["cbx_id"]}, {item["company"]}, {item["address"]}, {item["city"]}, {item["state"]} '
                       f'{item["country"]} {item["zip"]}, {item["email"]}, {item["first_name"]} {item["last_name"]}'
                       f' --> CR{item["ratio_company"]}, AR{item["ratio_address"]},'
                       f' CM{item["contact_match"]}, HCC{item["hiring_client_count"]}, M[{item["modules"]}]')
        # append matching results to the hc_list
        match_data = []
        uniques_cbx_id = set(item['cbx_id'] for item in matches)
        subscription_upgrade = False
        upgrade_price = 0.00
        prorated_upgrade_price = 0.00
        if uniques_cbx_id:
            for key, value in matches[0].items():
                # Skip matched_qstatus - it's only used for internal logic, not output
                if key != 'matched_qstatus':
                    match_data.append(value)
            hc_row.extend(match_data)
            hc_row.append(True if hc_domain in GENERIC_DOMAIN else False)
            hc_row.append(len(uniques_cbx_id) if len(uniques_cbx_id) else '')
            hc_row.append(len([i for i in matches if i['hiring_client_count'] > 0]))
            hc_row[HC_HEADER_LENGTH + analysis_headers.index("analysis")] = ('\n'.join(ids))
            # Calculate subscription upgrade and prorating
            if hc_row[HC_BASE_SUBSCRIPTION_FEE] == '':
                base_subscription_fee = CBX_DEFAULT_STANDARD_SUBSCRIPTION
                print(f'WARNING: no subscription fee defined for {hc_row[HC_COMPANY]}, using default {base_subscription_fee}')
            else:
                base_subscription_fee = hc_row[HC_BASE_SUBSCRIPTION_FEE]
            current_sub_total = matches[0]['subscription_price'] + matches[0]['employee_price']
            price_diff = float(base_subscription_fee) - current_sub_total
            if price_diff > 0 and matches[0]['registration_status'] == 'Active' and matches[0]['expiration_date'] \
                    and current_sub_total > 0.0:
                subscription_upgrade = True
                upgrade_price = price_diff
                expiration_date = matches[0]['expiration_date']
                now = datetime.now()
                if expiration_date > now:
                    delta = expiration_date - now
                    days = delta.days if delta.days < 365 else 365
                    prorated_upgrade_price = days / 365 * upgrade_price
                else:
                    prorated_upgrade_price = upgrade_price
                if smart_boolean(hc_row[HC_IS_ASSOCIATION_FEE]):
                    upgrade_price += 100.0
                    prorated_upgrade_price += 100
            if matches[0]['account_type'] in ('elearning', 'plan_nord', 'portail_pfr', 'special'):
                subscription_upgrade = True
                prorated_upgrade_price = upgrade_price = hc_row[HC_BASE_SUBSCRIPTION_FEE]
            # print(f'CBX Assessment Level: {matches[0]["cbx_assessment_level"]}.  Requested Assesssment Level: {parse_assessment_level(hc_row[HC_ASSESSMENT_LEVEL])}')
            if parse_assessment_level(matches[0]['cbx_assessment_level']) < parse_assessment_level(hc_row[HC_ASSESSMENT_LEVEL]):
                # print("Assessment level upgrade")
                subscription_upgrade = True
                prorated_upgrade_price = upgrade_price
        else:
            hc_row.extend(['' for x in range(len(analysis_headers)-6)])
        create_in_cognibox = False if len(uniques_cbx_id) and not hc_row[HC_AMBIGUOUS] else True
        hc_row.append(subscription_upgrade)
        hc_row.append(upgrade_price)
        hc_row.append(prorated_upgrade_price)
        hc_row.append(create_in_cognibox)
        hc_row.append(action(hc_row, matches[0] if len(matches) else {}, create_in_cognibox,
                             subscription_upgrade, matches[0]['expiration_date'] if len(matches) else None,
                             matches[0]['is_qualified'] if len(matches) else False, args.ignore_warnings))
        hc_row.append(index+1)
        metadata_array = []
        for md_index in metadata_indexes:
            metadata_array.insert(0, hc_row.pop(md_index))
        hc_row.extend(metadata_array)
        for i, value in enumerate(hc_row):
            out_ws.cell(index+2, i+1, value)
        if (index + 1) % 100 == 0:
            out_wb.save(filename=output_file)
        print(f'{index+1} of {total} [{len(uniques_cbx_id)} found]')

    out_wb.save(filename=output_file)

    hc_onboarding = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 'onboarding', hc_data)
    for index, row in enumerate(hc_onboarding):
        for i, value in enumerate(row):
            out_ws_onboarding.cell(index + 2, i + 1, value)

    hc_association_fee = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 'association_fee', hc_data)
    for index, row in enumerate(hc_association_fee):
        for i, value in enumerate(row):
            out_ws_association_fee.cell(index + 2, i + 1, value)

    hc_re_onboarding = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 're_onboarding', hc_data)
    for index, row in enumerate(hc_re_onboarding):
        for i, value in enumerate(row):
            out_ws_re_onboarding.cell(index + 2, i + 1, value)

    hc_subscription_upgrade = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 'subscription_upgrade',
                                     hc_data)
    for index, row in enumerate(hc_subscription_upgrade):
        for i, value in enumerate(row):
            out_ws_subscription_upgrade.cell(index + 2, i + 1, value)

    hc_ambiguous_onboarding = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 'ambiguous_onboarding',
                                     hc_data)
    for index, row in enumerate(hc_ambiguous_onboarding):
        for i, value in enumerate(row):
            out_ws_ambiguous_onboarding.cell(index + 2, i + 1, value)

    hc_restore_suspended = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 'restore_suspended',
                                  hc_data)
    for index, row in enumerate(hc_restore_suspended):
        for i, value in enumerate(row):
            out_ws_restore_suspended.cell(index + 2, i + 1, value)

    hc_activation_link = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 'activation_link',
                                hc_data)
    for index, row in enumerate(hc_activation_link):
        for i, value in enumerate(row):
            out_ws_activation_link.cell(index + 2, i + 1, value)

    hc_already_qualified = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 'already_qualified',
                                  hc_data)
    for index, row in enumerate(hc_already_qualified):
        for i, value in enumerate(row):
            out_ws_already_qualified.cell(index + 2, i + 1, value)

    hc_add_questionnaire = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 'add_questionnaire',
                                  hc_data)
    for index, row in enumerate(hc_add_questionnaire):
        for i, value in enumerate(row):
            out_ws_add_questionnaire.cell(index + 2, i + 1, value)

    hc_missing_information = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 'missing_info',
                                    hc_data)
    for index, row in enumerate(hc_missing_information):
        for i, value in enumerate(row):
            out_ws_missing_information.cell(index + 2, i + 1, value)

    hc_follow_up_qualification = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] ==
                                        'follow_up_qualification',
                                        hc_data)
    for index, row in enumerate(hc_follow_up_qualification):
        for i, value in enumerate(row):
            out_ws_follow_up_qualification.cell(index + 2, i + 1, value)

    existing_contractors_rd = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] != 'onboarding' and x[HC_HEADER_LENGTH+len(analysis_headers)-2] != 'missing_info' , hc_data)

    for index, row in enumerate(existing_contractors_rd):
        column = 0
        for i, value in enumerate(row):
            if i < len(existing_contractors_headers_mapping) and existing_contractors_headers_mapping[i]:
                column += 1
                out_ws_existing_contractors.cell(index + 2, column, value)

    hc_onboarding_rd = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 'onboarding',
                              hc_data)
    for index, row in enumerate(hc_onboarding_rd):
        column = 0
        for i, value in enumerate(row):
            if i < len(rd_headers_mapping) and rd_headers_mapping[i]:
                column += 1
                # Invert code and id columns
                if column == rd_pricing_group_id_col:
                    out_ws_onboarding_rd.cell(index + 2, column + 1, value)
                elif column == rd_pricing_group_code_col:
                    out_ws_onboarding_rd.cell(index + 2, column - 1, value)
                else:
                    out_ws_onboarding_rd.cell(index + 2, column, value)
    for index, row in enumerate(hc_data):
        column = 0
        for i, value in enumerate(row):
            if i < len(hs_headers_mapping) and hs_headers_mapping[i]:
                column += 1
                out_ws_onboarding_hs.cell(index + 2, column, value)

    # formatting the excel...
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    dims = {}
    for sheet in sheets:
        tab = Table(displayName=sheet.title.replace(" ", "_"),
                    ref=f'A1:{get_column_letter(sheet.max_column)}{sheet.max_row + 1}')
        tab.tableStyleInfo = style
        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            sheet.column_dimensions[col].width = value
        if sheet != out_ws_onboarding_rd:
            sheet.column_dimensions[get_column_letter(HC_HEADER_LENGTH+analysis_headers.index("hc_contractor_summary")+1)].width = 150
            sheet.column_dimensions[get_column_letter(HC_HEADER_LENGTH+analysis_headers.index("analysis")+1)].width = 150
            sheet.column_dimensions[get_column_letter(HC_HEADER_LENGTH+len(analysis_headers)-17)].width = 150
            sheet.column_dimensions[get_column_letter(HC_HEADER_LENGTH+len(analysis_headers)-18)].width = 150
            for i in range(2, len(hc_data)+1):
                sheet.cell(i, HC_HEADER_LENGTH+analysis_headers.index("analysis")+1).alignment = Alignment(wrapText=True)
                sheet.cell(i, HC_HEADER_LENGTH+analysis_headers.index("hc_contractor_summary")+1).alignment = Alignment(wrapText=True)
                sheet.cell(i, HC_HEADER_LENGTH+len(analysis_headers)-17).alignment = Alignment(wrapText=True)
                sheet.cell(i, HC_HEADER_LENGTH+len(analysis_headers)-18).alignment = Alignment(wrapText=True)
                
        sheet.add_table(tab)
    out_wb.save(filename=output_file)
    print(f'Completed data analysis...')
    print(f'Completed at {datetime.now()}')
